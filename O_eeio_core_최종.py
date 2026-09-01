"""
eeio_core_최종.py — 한국 EEIO 분석 공통 소스 모듈
=================================================
한국은행 투입산출표(IO Table) 기반 탄소 배출 분석 파이프라인의
핵심 함수 및 시각화 함수를 모두 이 파일에 정의합니다.

노트북(KR_EEIO_통합.ipynb)은 이 파일을 임포트하여 결과만 표시합니다.

포함 함수:
  - 데이터 로드    : load_ghg_data()
  - EEIO 계산      : run_eeio()  (경상/불변 공용)
  - 보조 함수      : _find_block, _pick_sheet, _total_input_vector,
                      clean_industry_names, clean_and_merge_pivot
  - 시각화 (공용)  : plot_heatmaps(), plot_bar_charts(), plot_timeseries()
  - 경상가격 전용  : build_scope_table_current(),
                      plot_scope_bar_donut_current(),
                      plot_dashboard_current()
  - 불변가격 전용  : build_scope_table_constant(),
                      plot_scope_bar_donut_constant(),
                      plot_facet_timeseries(), plot_total_emission_timeseries(),
                      plot_scope_share_timeseries()

  ★ [NEW] 위젯 없이 "연도만 지정"하면 대시보드 전체(인트로 + 전 산업 카드)를
     한 번에 순서대로 출력하는 함수 추가. 노트북을 껐다 켜도 위젯 상태가
     아니라 저장된 파일/코드 실행만으로 재현되므로 안정적입니다.

       - show_dashboard_year(year, p_type="경상")
             → 저장된 이미지(JSON 인덱스 기반)가 있으면 그걸 즉시 표시 (빠름)
       - show_dashboard_year_v2(year)
             → JSON 인덱스 없이 파일명 스캔만으로 동작하는 버전
       - show_dashboard_year_recompute(year, final_results, io_files, p_type=...)
             → 매번 새로 계산하여 표시 (해당 연도를 처음 볼 때 1회 사용)
       - show_dashboard_year_state(year, p_type="경상")
             → init_pipeline() 을 쓴 경우, 저장된 이미지 있으면 표시 /
               없으면 자동으로 재계산까지 알아서 수행하는 올인원 버전

     사용 예시 (노트북 셀에 아래처럼 연도만 바꿔서 실행):
         import eeio_core_최종 as ec
         ec.show_dashboard_year(2023, p_type="경상")
         ec.show_dashboard_year(2020, p_type="불변")

  ★★ [FIX 2026-08] run_eeio() 산업명 정규화 버그 수정 ★★
     기존에는 M_matrices에 저장되는 industry_names가 원본(raw, 정규화 전)
     그대로였고, 대시보드용 df_result의 산업명만 clean_industry_names()로
     정규화되었습니다. 이 때문에 원본 IO 파일에 "사회복지서비스"처럼
     띄어쓰기가 없는 표기가 남아있는 연도(예: 2015년)에서는
     plot_dashboard_with_scope() 등이 M_matrices['산업명'](raw)에서
     정규화된 이름(clean, 예: "사회복지 서비스")을 찾지 못해
     Upstream/Downstream이 "데이터 없음"으로 표시되는 문제가 있었습니다.
     아래 run_eeio()에서 industry_names를 만드는 시점에 바로
     clean_industry_names()를 적용하도록 수정하여, M_matrices/라벨/
     df_result가 모두 동일한(정규화된) 이름을 쓰도록 통일했습니다.
     ※ 이 수정 이후에는 해당 연도들에 대해 run_eeio() (또는 대시보드
        함수)를 다시 실행해 M_matrices 캐시를 갱신해야 반영됩니다.
"""

import os
import textwrap

import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
from matplotlib.patches import FancyBboxPatch
import seaborn as sns
from IPython.display import display, HTML

# ── 한글 폰트 설정 ──────────────────────────────────────────────────────────
import matplotlib.font_manager as fm

_candidates = [
    f.name for f in fm.fontManager.ttflist
    if any(k in f.name for k in ["Nanum", "Malgun", "AppleGothic", "Gulim", "NanumGothic"])
]
_font = _candidates[0] if _candidates else "DejaVu Sans"
plt.rcParams["font.family"] = _font
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["figure.facecolor"] = "white"
plt.rcParams["axes.facecolor"] = "white"

# 공유 M 행렬 저장소 (연도별)
M_matrices: dict = {}


# ═══════════════════════════════════════════════════════════════════════════════
# 1. 공통 유틸리티 함수
# ═══════════════════════════════════════════════════════════════════════════════

def _find_block(df: pd.DataFrame, code_col: int = 0):
    """코드열(A..T)에서 산업 블록의 시작/끝 행을 동적으로 찾는다."""
    codes = df.iloc[:, code_col].astype(str).str.strip().values
    start = int(np.where(codes == "A")[0][0])
    end   = int(np.where(codes == "T")[0][0])
    return start, end, codes


def _pick_sheet(xls: pd.ExcelFile, kw: str, p_type: str) -> str:
    """키워드+가격구분으로 시트명 선택 (언더스코어/공백 표기 혼용 대응)."""
    m = [s for s in xls.sheet_names if kw in s and p_type in s]
    if not m:
        m = [s for s in xls.sheet_names if kw in s]
    if not m:
        raise ValueError(f"'{kw}'({p_type}) 시트를 찾을 수 없습니다.")
    return m[0]


def _total_input_vector(df_tot: pd.DataFrame, size: int, code_col: int = 0) -> np.ndarray:
    """총거래표에서 '총투입계'(9790) 행 → 열별 소비산업 총투입 벡터."""
    rc = df_tot.iloc[:, code_col].astype(str).str.strip().values
    cand = np.where(rc == "9790")[0]
    if not len(cand):
        lbl  = df_tot.iloc[:, 1].astype(str).str.strip().values
        cand = np.where(lbl == "총투입계")[0]
    if not len(cand):
        raise ValueError("총거래표에서 '총투입계'(9790) 행을 찾지 못했습니다.")
    return df_tot.iloc[int(cand[0]), 2:2 + size].values.astype(float)


def clean_industry_names(idx: pd.Index) -> pd.Index:
    """산업명 인덱스의 띄어쓰기를 통일한다."""
    return (
        idx.astype(str).str.strip()
        .str.replace("복지서비스",            "복지 서비스",     regex=False)
        .str.replace("방송서비스",            "방송 서비스",     regex=False)
        .str.replace("보험서비스",            "보험 서비스",     regex=False)
        .str.replace(r"기타\s*제조업\s*제품", "기타 제조업 제품", regex=True)
        .str.replace("  ",                   " ",               regex=False)
    )


def clean_and_merge_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """피벗 테이블 산업명 정규화 및 중복 행 병합."""
    df = df.copy()
    df.index = clean_industry_names(df.index)
    return df.groupby(df.index).max()


def out_path(filename: str, output_dir: str = "./output") -> str:
    """output_dir 하위 경로를 반환하는 헬퍼 함수."""
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, filename)


# ═══════════════════════════════════════════════════════════════════════════════
# 2. 데이터 로드
# ═══════════════════════════════════════════════════════════════════════════════

def load_ghg_data(
    ghg_file: str = "2015-2023산업별_온실가스_추정_FF_송부.xlsx",
) -> pd.DataFrame:
    """
    산업별 온실가스 배출량 데이터 로드 (단위: kt CO₂eq).

    원본 마스터 빌더와 동일하게 header=1 로 읽어 '코드' 컬럼을 유지합니다.
    - 인덱스: 숫자형 (reset 상태)
    - '코드' 컬럼: 산업코드 (A, B, C01, ...)
    - '부문명' 또는 '산업명' 컬럼: 산업명
    - 연도 컬럼: '2015', '2016', ... (str)

    Returns
    -------
    df : pd.DataFrame  ('코드' 컬럼 포함, T 코드 제외)
    """
    df = pd.read_excel(ghg_file, sheet_name="연도별정리_최종", header=1)

    # 열 이름 정리: float → int → str (예: 2015.0 → '2015')
    new_cols = []
    for c in df.columns:
        try:
            new_cols.append(str(int(float(str(c)))))
        except (ValueError, TypeError):
            new_cols.append(str(c).strip())
    df.columns = new_cols

    # 코드열 정리
    df = df[df['코드'].notna()].copy()
    df['코드'] = df['코드'].astype(str).str.strip()

    # 산업명 컬럼 통일 ('부문명' → '산업명')
    if '부문명' in df.columns and '산업명' not in df.columns:
        df = df.rename(columns={'부문명': '산업명'})

    # 산업 코드: 알파벳으로 시작하는 행만 유지 (9090, 9111 등 집계/최종수요 행 제외)
    # 그 중 정확히 'T'(기타)인 행도 제외
    df = df[df['코드'].str.match(r'^[A-Za-z]', na=False)].copy()
    df = df[df['코드'] != 'T'].reset_index(drop=True)

    year_cols = [c for c in df.columns if c.isdigit() and 2000 <= int(c) <= 2100]
    print(f"✅ 온실가스 데이터 로드 완료  |  산업 수: {len(df)}개  |  연도: {year_cols}")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# 3. EEIO 계산 (경상/불변 공용)
# ═══════════════════════════════════════════════════════════════════════════════

def run_eeio(
    year,
    file_name: str,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
    exclude_keyword: str | None = None,
    ghg_unit_scale: float = 1000.0,
) -> pd.DataFrame | None:
    """
    국산 기술계수행렬 기반 EEIO 계수(직접·공급망·Scope1/2/3) 계산.

    Parameters
    ----------
    year             : 분석 연도 (int 또는 str)
    file_name        : 해당 연도 IO 엑셀 파일 경로
    ghg_df           : load_ghg_data() 반환값 (kt CO₂eq)
    p_type           : '경상' 또는 '불변'
    exclude_keyword  : 제외할 산업명 prefix (예: '기타')
                       ※ 정확히 '기타'로 시작하는 행만 제외 (기타서비스 등 포함 X)
    ghg_unit_scale   : kt → t 변환 배율 (기본 1000)

    Returns
    -------
    df_result : pd.DataFrame, 인덱스=산업명
        컬럼: 직접배출계수(B), 공급망유발계수(B*L), 간접유발량(BL-B), Scope2, Scope3
    """
    try:
        xls = pd.ExcelFile(file_name)

        tot_sheet = _pick_sheet(xls, "총거래표",  p_type)
        dom_sheet = _pick_sheet(xls, "국산거래표", p_type)
        df_tot = pd.read_excel(file_name, sheet_name=tot_sheet, header=None)
        df_dom = pd.read_excel(file_name, sheet_name=dom_sheet, header=None)

        s_t, e_t, codes_t = _find_block(df_tot)
        s_d, e_d, codes_d = _find_block(df_dom)
        sector_codes   = codes_d[s_d:e_d + 1]
        if not np.array_equal(codes_t[s_t:e_t + 1], sector_codes):
            raise ValueError("총거래표와 국산거래표의 산업 코드 순서가 다릅니다.")

        # ★ [FIX] 산업명을 여기서 바로 정규화한다 (기존 버그: 아래에서 만드는
        #   industry_names 가 raw 그대로 M_matrices/라벨에 저장되어, 대시보드가
        #   참조하는 clean(정규화)된 이름과 표기가 달라 매칭에 실패하는
        #   문제(예: 2015년 "사회복지서비스" vs "사회복지 서비스")가 있었다.
        #   이렇게 소스에서부터 정규화하면 M_matrices['산업명']/['라벨']과
        #   df_result 의 인덱스가 항상 동일한 표기를 쓰게 되어 일관성이
        #   보장된다. df_result 에 대한 기존의 clean_industry_names() 호출은
        #   이제 이미 정규화된 값을 다시 정규화하는 것이라 무해하다(idempotent).
        industry_names = clean_industry_names(
            pd.Index(df_dom.iloc[s_d:e_d + 1, 1].values.astype(str))
        ).values
        size = len(sector_codes)

        idx_D = np.where(sector_codes == "D")[0]
        if not len(idx_D):
            raise ValueError("전력(D) 코드가 없습니다.")
        D_rel = int(idx_D[0])

        X  = _total_input_vector(df_tot, size)
        Xs = np.where((X == 0) | np.isnan(X), np.nan, X)

        Z = df_dom.iloc[s_d:e_d + 1, 2:2 + size].values.astype(float)
        A = np.nan_to_num(Z / Xs.reshape(1, -1), nan=0.0, posinf=0.0, neginf=0.0)
        L = np.linalg.inv(np.eye(size) - A)

        # GHG: '코드' 컬럼을 인덱스로 세팅한 뒤 연도 컬럼 추출
        # ★ 중요: kt -> t 변환(ghg_unit_scale, 기본 1000배)을 B/M/M_X 계산 이전에
        #   적용한다. 예전 버전은 이 변환을 df_result 표시 단계에서만 곱해서,
        #   M_matrices 에 저장되는 M/B/BL/M_X 가 kt 단위 그대로 남아있는 버그가
        #   있었다 (즉 대시보드/엑셀에 쓰이는 실제 값이 1000배 작았음).
        _yr_key = str(year)
        if _yr_key not in ghg_df.columns:
            raise KeyError(f"온실가스 데이터에 {year}년 컬럼이 없습니다. (보유: {list(ghg_df.columns)})")
        E_kt = (
            ghg_df.set_index('코드')[_yr_key]
            .reindex(pd.Index(sector_codes.tolist()))
            .fillna(0)
            .values
            .astype(float)
        )
        E = E_kt * ghg_unit_scale   # kt -> t 변환 (여기서 바로 적용)
        B  = np.nan_to_num(E / Xs, nan=0.0, posinf=0.0, neginf=0.0)
        BL = (B.reshape(1, -1) @ L).flatten()
        M  = np.diag(B) @ L

        # ── M_X: 금액 유발행렬 (M x diag(X)) ──────────────────────────────
        # M[i,j] = j산업이 1백만원 생산할 때 i산업 경유로 유발되는 배출계수
        # (원단위, t CO2eq / 백만원). 여기에 j산업의 실제 총투입액 X[j]를
        # 곱하면 j산업의 실제 생산 규모에서 i산업 경유로 실제 유발된
        # 절대 배출량(t CO2eq)이 된다. 대시보드 카드의 Upstream/Downstream
        # 표와 엑셀 M_X 시트가 동일한 값을 참조하도록 통일한다.
        X_for_scale = np.nan_to_num(X, nan=0.0)
        M_X = M @ np.diag(X_for_scale)

        # ── 코드+이름 MultiIndex 라벨 (행: 코드행+이름행, 열: 코드열+이름열로
        #    분리 표시하기 위한 (코드, 이름) 튜플 배열) ────────────────────
        labels = pd.MultiIndex.from_arrays(
            [sector_codes.tolist(), industry_names.tolist()],
            names=["코드", "산업명"],
        )

        # ★ 키에 연도뿐 아니라 p_type(경상/불변)도 포함시킨다. 예전에는 키가
        #   연도(year)만이라, 같은 연도를 경상→불변(또는 그 반대) 순서로
        #   반복 계산하면 나중 것이 앞의 것을 덮어써서, 이후 build_eeio_matrices()가
        #   "이미 계산돼 있으니 재사용"한다며 엉뚱한 가격기준의 값으로 엑셀을
        #   만들어버리는 버그가 있었다 (대시보드 Upstream/Downstream과 엑셀
        #   M.X(diag) 값이 서로 달라지는 근본 원인).
        mkey = f"{year}_{p_type}"
        M_matrices[mkey] = {
            "M": M.copy(), "M_X": M_X.copy(), "L": L.copy(), "A": A.copy(),
            "B": B.copy(), "BL": BL.copy(), "X": X.copy(), "GHG_raw": E.copy(),
            "산업명": industry_names.copy(), "코드": sector_codes.copy(),
            "라벨": labels,
        }
        # 하위 호환: 기존에 "연도만" 키로 접근하던 코드가 있다면 최신 계산
        # 결과를 계속 볼 수 있도록 연도 키도 함께 갱신해 둔다(참고용).
        M_matrices[year] = M_matrices[mkey]

        # Scope 1 / 2 / 3 (이미 t 단위인 B/M/BL 기준으로 계산 — 추가 스케일 불필요)
        scope_1     = B.copy()
        scope_2_raw = M[D_rel, :].copy()
        scope_2_raw[D_rel] -= scope_1[D_rel]
        scope_2 = np.maximum(scope_2_raw, 0.0)

        scope_3_raw = BL - scope_1 - scope_2
        scope_3 = np.maximum(scope_3_raw, 0.0)

        neg_n = int((scope_3_raw < -1e-9).sum())
        if neg_n:
            print(f"⚠️ {year}: Scope3 음수 {neg_n}개 (클립으로 손실). "
                  f"min={scope_3_raw.min():.4g} — 정의 점검 권장")

        df_result = pd.DataFrame({
            "산업명":             industry_names,
            "직접배출계수(B)":     scope_1,
            "공급망유발계수(B*L)": BL,
            "간접유발량(BL-B)":   (BL - scope_1),
            "Scope2":            scope_2,
            "Scope3":            scope_3,
        })

        # ── 정확히 '기타'로만 시작하는 산업만 제외 ──────────────────────────
        # str.startswith('기타') 는 '기타서비스', '기타금융' 등 모두 포함하므로
        # 정확히 산업명 == '기타' 이거나 '기타 ' 로 시작하는 경우만 제외한다.
        if exclude_keyword:
            mask = df_result["산업명"].astype(str).str.strip() == exclude_keyword
            df_result = df_result[~mask].copy()

        df_result["산업명"] = clean_industry_names(pd.Index(df_result["산업명"].astype(str)))
        return df_result.set_index("산업명")

    except Exception as e:
        print(f"❌ {year}년 오류 발생: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# 4. 전체 연도 일괄 실행
# ═══════════════════════════════════════════════════════════════════════════════

def run_all_years(
    io_files: dict,
    analysis_years: list,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
    exclude_keyword: str | None = None,
    output_dir: str = "./output",
) -> dict:
    """전체 연도 EEIO 분석 실행 → final_results dict 반환."""
    final_results = {}
    label = "경상가격" if p_type == "경상" else "불변가격"
    print("━" * 70)
    print(f"  전체 연도 EEIO 분석 ({label})")
    print("━" * 70)

    for yr, fpath in io_files.items():
        res = run_eeio(yr, fpath, ghg_df, p_type=p_type, exclude_keyword=exclude_keyword)
        if res is None:
            continue
        final_results[yr] = res
        res_sorted = res.sort_values("공급망유발계수(B*L)", ascending=False)
        csv_file   = out_path(f"Korea_EEIO_{p_type}_{yr}.csv", output_dir)
        res_sorted.to_csv(csv_file, encoding="utf-8-sig")
        print(f"    ✅ {yr}년 완료 ({len(res_sorted)}개 산업) → {csv_file}")

    print("━" * 70)
    print(f"  분석 완료 — 총 {len(final_results)}개 연도")
    print("━" * 70)
    return final_results


# ═══════════════════════════════════════════════════════════════════════════════
# 5. 시각화용 데이터 통합
# ═══════════════════════════════════════════════════════════════════════════════

def build_viz_data(final_results: dict):
    """
    final_results → (total_viz, pivot_BL, pivot_B) 반환.

    Returns
    -------
    total_viz : 롱 포맷 DataFrame
    pivot_BL  : 공급망유발계수 피벗 (산업 × 연도)
    pivot_B   : 직접배출계수 피벗   (산업 × 연도)
    """
    frames = []
    for yr, df in final_results.items():
        tmp = df.copy().reset_index()
        tmp["연도"] = int(yr)
        frames.append(tmp)

    total_viz = pd.concat(frames, ignore_index=True)

    pivot_BL = clean_and_merge_pivot(
        total_viz.pivot(index="산업명", columns="연도", values="공급망유발계수(B*L)")
    )
    pivot_B = clean_and_merge_pivot(
        total_viz.pivot(index="산업명", columns="연도", values="직접배출계수(B)")
    )

    order_2023 = pivot_BL[2023].sort_values(ascending=False).index
    pivot_BL   = pivot_BL.loc[order_2023]
    pivot_B    = pivot_B.loc[order_2023]

    print(f"✅ 통합 완료 — {len(total_viz['산업명'].unique())}개 산업 × {len(total_viz['연도'].unique())}개 연도")
    return total_viz, pivot_BL, pivot_B


# ═══════════════════════════════════════════════════════════════════════════════
# 6. 공통 시각화 함수
# ═══════════════════════════════════════════════════════════════════════════════

def plot_heatmaps(pivot_BL, pivot_B, p_type: str = "경상", output_dir: str = "./output"):
    """연도 × 산업 히트맵 (공급망 유발계수 + 직접 배출계수)."""
    pivots    = [pivot_BL, pivot_B]
    titles    = ["공급망 유발계수 (B×L)", "직접 배출계수 (B)"]
    cmaps     = ["YlGn", "Blues"]
    filenames = [
        out_path(f"Heatmap_{p_type}_공급망유발계수.png", output_dir),
        out_path(f"Heatmap_{p_type}_직접배출계수.png",   output_dir),
    ]
    for pivot, title, cmap, filename in zip(pivots, titles, cmaps, filenames):
        plt.figure(figsize=(14, 16), facecolor="white")
        sns.heatmap(
            pivot, cmap=cmap, annot=True, fmt=".4f",
            linewidths=0.4, linecolor="#e5e7eb",
            cbar_kws={"label": "t CO₂eq / 백만원", "shrink": 0.8},
            annot_kws={"size": 10},
        )
        plt.title(f"한국 산업별 {title}\n({p_type}가격)", fontsize=18, weight="bold", pad=20)
        plt.xlabel("연도", fontsize=13, labelpad=12)
        plt.ylabel("산업 부문", fontsize=13, labelpad=12)
        plt.xticks(rotation=0, fontsize=12)
        plt.yticks(rotation=0, fontsize=11)
        plt.tight_layout()
        plt.savefig(filename, dpi=200, bbox_inches="tight")
        plt.show()
        print(f"🖼️  저장 → {filename}\n")


def plot_bar_charts(total_viz, analysis_years: list, p_type: str = "경상", output_dir: str = "./output"):
    """연도별 전 산업 직접 vs 공급망 비교 가로 막대 차트."""
    for yr in analysis_years:
        if yr not in total_viz["연도"].unique():
            continue
        df_yr = total_viz[total_viz["연도"] == yr].sort_values("공급망유발계수(B*L)", ascending=False)
        df_m  = df_yr.melt(
            id_vars="산업명",
            value_vars=["직접배출계수(B)", "공급망유발계수(B*L)"],
            var_name="구분", value_name="계수",
        ).replace({
            "직접배출계수(B)":     "① 직접 배출 (B)",
            "공급망유발계수(B*L)": "② 공급망 포함 총유발 (B×L)",
        })
        n   = len(df_yr)
        fig, ax = plt.subplots(figsize=(13, max(6, n * 0.42)), facecolor="white")
        palette = {"① 직접 배출 (B)": "#74C69D", "② 공급망 포함 총유발 (B×L)": "#1B4332"}
        sns.barplot(data=df_m, x="계수", y="산업명", hue="구분",
                    palette=palette, ax=ax, orient="h")
        for bar in ax.patches:
            w = bar.get_width()
            if w > 0:
                ax.text(w + ax.get_xlim()[1] * 0.005, bar.get_y() + bar.get_height() / 2,
                        f"{w:.4f}", va="center", ha="left", fontsize=7, color="#374151")
        ax.set_title(f"{yr}년 전 산업 직접 vs 공급망 탄소 유발계수 ({p_type}가격)",
                     fontsize=14, pad=14, weight="bold")
        ax.set_xlabel("탄소 유발계수 (t CO₂eq / 백만원)", fontsize=11, labelpad=8)
        ax.set_ylabel("")
        ax.grid(axis="x", linestyle="--", alpha=0.4)
        ax.legend(title="구분", fontsize=10, title_fontsize=11, loc="lower right", framealpha=0.9)
        plt.tight_layout()
        img = out_path(f"Bar_{p_type}_{yr}.png", output_dir)
        plt.savefig(img, dpi=150, bbox_inches="tight")
        plt.show()
        print(f"   🖼️  {yr}년 저장 → {img}")


def plot_timeseries(total_viz, pivot_BL, pivot_B, analysis_years: list,
                    p_type: str = "경상", output_dir: str = "./output"):
    """주요 산업 시계열 라인 차트 (공급망 유발계수 상위 8 + 직접 배출계수 상위 8)."""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(15, 16), facecolor="white")
    palette8 = sns.color_palette("Dark2", 8)

    top8   = pivot_BL[2023].nlargest(8).index.tolist()
    df_ts  = total_viz[total_viz["산업명"].isin(top8)].sort_values("연도")
    for i, sec in enumerate(top8):
        sub = df_ts[df_ts["산업명"] == sec]
        ax1.plot(sub["연도"], sub["공급망유발계수(B*L)"],
                 marker="o", ms=8, lw=2.5, label=sec, color=palette8[i])
    ax1.set_title(f"공급망 유발계수 상위 8개 산업 시계열 ({p_type}가격)",
                  fontsize=14, pad=12, weight="bold")
    ax1.set_ylabel("공급망 유발계수 (t CO₂eq / 백만원)", fontsize=11)
    ax1.set_xticks(analysis_years)
    ax1.set_xticklabels([str(y) for y in analysis_years])
    ax1.grid(True, linestyle="--", alpha=0.4)
    ax1.legend(title="산업", fontsize=9, title_fontsize=10, loc="upper right", framealpha=0.9)

    bottom8 = pivot_B[2023].nlargest(8).index.tolist()
    df_ts2  = total_viz[total_viz["산업명"].isin(bottom8)].sort_values("연도")
    for i, sec in enumerate(bottom8):
        sub = df_ts2[df_ts2["산업명"] == sec]
        ax2.plot(sub["연도"], sub["직접배출계수(B)"],
                 marker="s", ms=7, lw=2.5, linestyle="--", label=sec, color=palette8[i])
    ax2.set_title(f"직접 배출계수 상위 8개 산업 시계열 ({p_type}가격)",
                  fontsize=14, pad=12, weight="bold")
    ax2.set_xlabel("연도", fontsize=11, labelpad=8)
    ax2.set_ylabel("직접 배출계수 (t CO₂eq / 백만원)", fontsize=11)
    ax2.set_xticks(analysis_years)
    ax2.set_xticklabels([str(y) for y in analysis_years])
    ax2.grid(True, linestyle="--", alpha=0.4)
    ax2.legend(title="산업", fontsize=9, title_fontsize=10, loc="upper right", framealpha=0.9)

    plt.tight_layout()
    fpath = out_path(f"Timeseries_{p_type}.png", output_dir)
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"🖼️  저장 → {fpath}")


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Scope 테이블 생성 (경상/불변 공용)
# ═══════════════════════════════════════════════════════════════════════════════

def build_scope_table(
    final_results: dict,
    io_files: dict,
    p_type: str = "경상",
    target_year: str = "2023",
    exclude_keyword: str | None = None,
) -> pd.DataFrame:
    """
    target_year 기준 Scope 1/2/3 배출량 테이블 생성.

    ※ 산업명이 정확히 exclude_keyword 인 행만 제외.
    """
    try:
        xls        = pd.ExcelFile(io_files[target_year])
        tot_sheet  = _pick_sheet(xls, "총거래표", p_type)
        df_tot     = pd.read_excel(io_files[target_year], sheet_name=tot_sheet, header=None)
        s_d, e_d, _ = _find_block(df_tot)
        names_raw  = df_tot.iloc[s_d:e_d + 1, 1].values
        vals       = _total_input_vector(df_tot, len(names_raw))
        univ_inp   = dict(zip(names_raw, vals))
        print(f"✅ {target_year}년 투입 구조 로드 완료 ({p_type}가격)")
    except Exception as e:
        print(f"❌ 로드 실패: {e} → 균등 더미값 사용")
        univ_inp = {k: 100.0 for k in final_results[target_year].index}

    res     = final_results[target_year]
    b_dict  = res["직접배출계수(B)"].to_dict()
    s2_dict = res["Scope2"].to_dict()
    s3_dict = res["Scope3"].to_dict()
    bl_dict = res["공급망유발계수(B*L)"].to_dict()

    rows = []
    for sec, val in univ_inp.items():
        sec_clean = str(sec).strip()
        if exclude_keyword and sec_clean == exclude_keyword:
            continue
        if sec_clean not in b_dict or val <= 0:
            continue
        rows.append({
            "산업":     sec_clean,
            "투입가치": val,
            "Scope1":   val * b_dict[sec_clean],
            "Scope2":   val * s2_dict[sec_clean],
            "Scope3":   val * s3_dict[sec_clean],
            "합계":     val * bl_dict[sec_clean],
            "탄소집약도": bl_dict[sec_clean],
        })

    return pd.DataFrame(rows).sort_values("합계", ascending=False)


# ═══════════════════════════════════════════════════════════════════════════════
# 8. 경상가격 전용 — Scope 도넛/바 차트 (연도 선택형)
# ═══════════════════════════════════════════════════════════════════════════════

def plot_dashboard_with_scope(
    final_results: dict,
    io_files: dict,
    p_type: str,
    year: str,
    exclude_keyword: str | None = None,
    output_dir: str = "./output",
) -> None:
    """
    원본 경상가격 STEP 11 — 산업별 Scope 1/2/3 종합 대시보드 (EEIO-IDA 스타일).
    exclude_keyword 와 정확히 일치하는 산업명만 제외 (startswith 아님).

    구성:
      1) 종합 인트로 표지 (전체 인사이트)
      2) 산업별 카드 순차 출력 (도넛 + Upstream/Downstream + 설명)

    ※ 이 함수를 1회 실행하면 인트로/카드 PNG와 Dashboard_Index_{p_type}_{year}.json
      이 output_dir 에 저장되어, 이후에는 show_dashboard_year() 로 재계산 없이
      즉시 재표시할 수 있습니다.
    """
    import textwrap
    import matplotlib as mpl
    from matplotlib.patches import FancyBboxPatch

    mpl.rcParams['figure.dpi'] = 200

    # ── df_scope 빌드 ────────────────────────────────────────────────────────
    try:
        xls       = pd.ExcelFile(io_files[year])
        tot_sheet = _pick_sheet(xls, '총거래표', p_type)
        df_tot    = pd.read_excel(io_files[year], sheet_name=tot_sheet, header=None)
        s_d, e_d, _ = _find_block(df_tot)
        names_raw = df_tot.iloc[s_d:e_d + 1, 1].values
        vals_raw  = _total_input_vector(df_tot, len(names_raw))
        # run_eeio() 결과(b_dict 등)는 clean_industry_names() 로 정규화된 이름을
        # 인덱스로 쓰므로, 여기서도 동일하게 정규화해야 매칭이 깨지지 않는다.
        # (예: 2020년 원본 '방송서비스' 처럼 공백이 없는 표기가 남아있으면
        #  정규화된 이름과 달라져 해당 산업이 조용히 스킵된다.)
        names_clean = clean_industry_names(pd.Index(names_raw.astype(str))).values
        univ_inp  = dict(zip(names_clean, vals_raw))
    except Exception as e:
        print(f"❌ 로드 실패: {e} → 균등 더미값 사용")
        univ_inp = {k: 100.0 for k in final_results[year].index}

    res     = final_results[year]
    b_dict  = res['직접배출계수(B)'].to_dict()
    s2_dict = res['Scope2'].to_dict()
    s3_dict = res['Scope3'].to_dict()
    bl_dict = res['공급망유발계수(B*L)'].to_dict()

    scope_rows = []
    for sec, val in univ_inp.items():
        sec_s = str(sec).strip()
        if exclude_keyword and sec_s == exclude_keyword:   # 정확히 일치만 제외
            continue
        if sec_s not in b_dict or val <= 0:
            continue
        scope_rows.append({
            '산업':     sec_s,
            '투입가치': val,
            'Scope1':   val * b_dict[sec_s],
            'Scope2':   val * s2_dict[sec_s],
            'Scope3':   val * s3_dict[sec_s],
            '합계':     val * bl_dict[sec_s],
            '탄소집약도': val * bl_dict[sec_s] / val,
        })

    df_scope = pd.DataFrame(scope_rows).sort_values('합계', ascending=False)

    # ── 대시보드 파라미터 ────────────────────────────────────────────────────
    DASH_YEAR     = year
    TOP_N_CONTRIB = 5
    MY_COLORS     = ['#1B4332', '#2D6A4F', '#74C69D']
    label         = '경상가격' if p_type == '경상' else '불변가격'

    df_dash   = df_scope.copy().reset_index(drop=True)
    n_sectors = len(df_dash)

    # ★ 키에 p_type 을 포함해 조회한다 (run_eeio() 가 저장한 것과 동일 규칙).
    #   이렇게 해야 같은 연도를 경상/불변으로 번갈아 계산해도 서로 덮어쓰지
    #   않고, 지금 그리는 p_type 에 정확히 대응하는 M/M_X 를 사용한다.
    _mkey = f"{DASH_YEAR}_{p_type}"
    if _mkey in M_matrices:
        mdat = M_matrices[_mkey]
    else:
        # 혹시 이 정확한 키가 없다면(예: 구버전 캐시) 마지막 계산 결과로 대체
        mdat = M_matrices.get(DASH_YEAR, list(M_matrices.values())[-1])
    sector_names_M = list(mdat['산업명'])
    # ── Upstream/Downstream 은 엑셀 M_X 시트(금액 유발행렬, 단위: t CO2eq)와
    #    동일한 값을 쓴다. M_X = M(diag(B)*L, 원단위) x diag(X)(실제 총투입액).
    #    kt -> t 변환(x1000)은 run_eeio() 안에서 이미 B 계산 시 반영되어 있으므로
    #    여기서는 추가로 곱하지 않는다 (예전 버전은 M*1000 을 잘못 곱해
    #    엑셀의 M_X 와 값이 어긋났었다).
    M_X_mat        = mdat['M_X']

    def get_M_index(name):
        try:    return sector_names_M.index(name)
        except ValueError: return None

    def draw_sector_card(fig, gs_row, sector_row, rank):
        sec = sector_row['산업']
        scope1, scope2, scope3 = sector_row['Scope1'], sector_row['Scope2'], sector_row['Scope3']
        total_12  = scope1 + scope2
        total_123 = scope1 + scope2 + scope3
        if total_123 <= 0:
            total_123 = 1e-12

        idx = get_M_index(sec)
        if idx is not None:
            # col: idx 산업(j) 이 생산될 때, 각 산업(i)을 "경유"해 유발되는
            #      절대 배출량(t CO2eq) → "Upstream(Backward)": idx 산업의
            #      생산을 위해 어느 산업의 배출이 유발되는가
            col      = M_X_mat[:, idx].copy(); col[idx] = 0
            up_idx   = np.argsort(col)[::-1][:TOP_N_CONTRIB]
            upstream_top = [(sector_names_M[i], col[i]) for i in up_idx]
            # rowv: idx 산업(i) 의 배출이 각 산업(j)의 생산에 의해 얼마나
            #      유발되었는가 → "Downstream(Forward)"
            rowv     = M_X_mat[idx, :].copy(); rowv[idx] = 0
            dn_idx   = np.argsort(rowv)[::-1][:TOP_N_CONTRIB]
            downstream_top = [(sector_names_M[i], rowv[i]) for i in dn_idx]
        else:
            upstream_top = downstream_top = []

        gs_card  = gs_row.subgridspec(1, 4, width_ratios=[1.15, 0.20, 0.9, 1.45], wspace=0.05)

        # 좌측 패널
        ax_left = fig.add_subplot(gs_card[0, 0])
        ax_left.axis('off')
        ax_left.set_xlim(0, 1); ax_left.set_ylim(-0.6, 1.15)

        title_text    = f"#{rank}  {sec}"
        title_fontsize = 13 if len(title_text) <= 13 else (11.5 if len(title_text) <= 20 else 10.5)
        ax_left.text(0, 1.10, title_text, fontsize=title_fontsize, weight='bold', color='#16a34a', va='top')
        ax_left.text(0, 0.99, f"한국 EEIO 분석 결과 | {DASH_YEAR}년 {label}", fontsize=8.5, color='#6b7280', va='top')

        box_y = 0.88
        ax_left.add_patch(FancyBboxPatch((0, box_y-0.085), 0.46, 0.11, boxstyle="round,pad=0.012",
                           facecolor='#f3f4f6', edgecolor='#d1d5db', transform=ax_left.transData, clip_on=False))
        ax_left.text(0.23, box_y-0.012, "Scope 1+2", fontsize=8.5, ha='center', color='#374151', weight='bold')
        ax_left.text(0.23, box_y-0.065, f"{total_12:,.4f}", fontsize=12, ha='center', color='#111827', weight='bold')

        ax_left.add_patch(FancyBboxPatch((0.50, box_y-0.085), 0.46, 0.11, boxstyle="round,pad=0.012",
                           facecolor='#dcfce7', edgecolor='#16a34a', transform=ax_left.transData, clip_on=False))
        ax_left.text(0.73, box_y-0.012, "Scope 1+2+3 (총배출)", fontsize=8.5, ha='center', color='#15803d', weight='bold')
        ax_left.text(0.73, box_y-0.065, f"{total_123:,.4f}", fontsize=12, ha='center', color='#15803d', weight='bold')

        ax_left.text(0, box_y-0.135, f"단위: t CO₂eq / 백만원 ({label} 기준)", fontsize=8, color='#9ca3af')

        line_gap = 0.085
        uy = box_y - 0.23
        ax_left.text(0, uy, "Top 5 Upstream  (이 산업 생산으로 실제 유발된 타 산업 배출량, t CO2eq)", fontsize=9, weight='bold', color='#1B4332')
        for i, (nm, val) in enumerate(upstream_top or [("데이터 없음", 0)]):
            ax_left.text(0.02, uy - line_gap*(i+1), f"{i+1}. {nm}", fontsize=8.5, color='#374151')
            ax_left.text(0.98, uy - line_gap*(i+1), f"{val:,.4f}",   fontsize=8.5, color='#6b7280', ha='right')

        dy = uy - line_gap * 6.2
        ax_left.text(0, dy, "Top 5 Downstream  (이 산업 배출이 타 산업 생산에 실제 유발한 양, t CO2eq)", fontsize=9, weight='bold', color='#2D6A4F')
        for i, (nm, val) in enumerate(downstream_top or [("데이터 없음", 0)]):
            ax_left.text(0.02, dy - line_gap*(i+1), f"{i+1}. {nm}", fontsize=8.5, color='#374151')
            ax_left.text(0.98, dy - line_gap*(i+1), f"{val:,.4f}",   fontsize=8.5, color='#6b7280', ha='right')

        # 중앙 패널: 도넛
        ax_donut = fig.add_subplot(gs_card[0, 2])
        vals = [max(scope1, 0), max(scope2, 0), max(scope3, 0)]
        if sum(vals) <= 0:
            vals = [1, 0, 0]
        wedges, _texts = ax_donut.pie(
            vals, colors=MY_COLORS, startangle=90,
            radius=0.85,
            wedgeprops={'width': 0.36, 'edgecolor': 'white', 'linewidth': 1.5}
        )
        # 외부 callout: Scope명 + % 수치를 조각 색상으로 표시
        _slabels = ['Scope 1\n(직접)', 'Scope 2\n(에너지)', 'Scope 3\n(공급망)']
        total_v  = sum(vals) or 1e-12
        for wedge, v, slbl in zip(wedges, vals, _slabels):
            pct = v / total_v * 100
            if pct < 0.01:
                continue  # 사실상 0인 경우만 생략
            ang = (wedge.theta1 + wedge.theta2) / 2
            rad = np.deg2rad(ang)
            r_ring = 0.85 * (1 - 0.36 / 2)   # 고리 중심 반경
            r_tip  = 0.90                      # 선 시작점
            r_bend = 1.05                      # 꺾임점
            r_text = 1.12                      # 텍스트
            _xt, _yt = r_text * np.cos(rad), r_text * np.sin(rad)
            _xb, _yb = r_bend * np.cos(rad), r_bend * np.sin(rad)
            _xs, _ys = r_tip  * np.cos(rad), r_tip  * np.sin(rad)
            # 연결선 (고리 바깥 끝 → 꺾임점)
            ax_donut.plot([_xs, _xb], [_ys, _yb],
                          color=wedge.get_facecolor(), lw=1.0, solid_capstyle='round')
            ha = 'left' if _xt >= 0 else 'right'
            ax_donut.text(
                _xt, _yt,
                f"{slbl}\n{pct:.1f}%",
                ha=ha, va='center', fontsize=7.5, weight='bold',
                color=wedge.get_facecolor(),
                linespacing=1.3,
            )
        ax_donut.set_xlim(-1.7, 1.7); ax_donut.set_ylim(-1.7, 1.7)
        sec_wrapped   = textwrap.fill(sec, width=10)
        name_fontsize = 12 if len(sec) <= 6 else (10.5 if len(sec) <= 12 else 9.5)
        ax_donut.text(0, 0, sec_wrapped, ha='center', va='center',
                      fontsize=name_fontsize, weight='bold', linespacing=1.25)

        # 우측 패널: 설명
        ax_right = fig.add_subplot(gs_card[0, 3])
        ax_right.axis('off')
        ax_right.set_xlim(0, 1.25); ax_right.set_ylim(-0.4, 1.15)
        rx = 0.12
        ax_right.text(rx, 1.10, "Scope 1, 2, 3 배출 구조", fontsize=11, weight='bold', color='#111827', va='top')
        sec_right_wrapped = textwrap.fill(sec, width=22)
        ax_right.text(rx, 0.98, sec_right_wrapped, fontsize=10, weight='bold', color='#16a34a', va='top', linespacing=1.3)

        n_right_lines = sec_right_wrapped.count('\n') + 1
        s1p = scope1 / total_123 * 100
        s2p = scope2 / total_123 * 100
        s3p = scope3 / total_123 * 100
        ty  = 0.82 - 0.08 * max(n_right_lines - 1, 0)
        gap_y = 0.30

        ax_right.text(rx, ty,      f"Scope 1 ({s1p:.0f}%)", fontsize=9.5, weight='bold', color=MY_COLORS[0])
        ax_right.text(rx, ty-0.08, f"직접 연소, 공정 배출이 이 산업\n배출량의 {s1p:.0f}%를 차지합니다.",
                      fontsize=9, color='#374151', va='top', linespacing=1.4)
        ty2 = ty - gap_y
        ax_right.text(rx, ty2,      f"Scope 2 ({s2p:.0f}%)", fontsize=9.5, weight='bold', color=MY_COLORS[1])
        ax_right.text(rx, ty2-0.08, f"구매 전력, 가스, 증기 사용에서\n발생하는 간접 배출은 {s2p:.0f}%입니다.",
                      fontsize=9, color='#374151', va='top', linespacing=1.4)
        ty3 = ty2 - gap_y
        ax_right.text(rx, ty3,      f"Scope 3 ({s3p:.0f}%)", fontsize=9.5, weight='bold', color=MY_COLORS[2])
        ax_right.text(rx, ty3-0.08, f"나머지 {s3p:.0f}%는 공급망 전체\n(원자재, 운송 등)에서 유발됩니다.",
                      fontsize=9, color='#374151', va='top', linespacing=1.4)
        ax_right.text(rx, -0.30, f"* 산업 집계 데이터 기반 근사치, {label} 기준",
                      fontsize=7.5, color='#9ca3af', va='bottom')

    # ── 1. 인트로 표지 ───────────────────────────────────────────────────────
    total_all  = df_dash['합계'].sum() or 1e-12
    s1a, s2a, s3a = df_dash['Scope1'].sum(), df_dash['Scope2'].sum(), df_dash['Scope3'].sum()
    top3       = df_dash.head(3)
    top3_share = top3['합계'].sum() / total_all * 100
    s1_sh = (df_dash['Scope1'] / df_dash['합계'].replace(0, np.nan) * 100).fillna(0)
    s3_sh = (df_dash['Scope3'] / df_dash['합계'].replace(0, np.nan) * 100).fillna(0)

    insight_lines = [
        f"전체 {n_sectors}개 산업의 합산 탄소 발자국은 {total_all:,.4f} t CO₂eq이며, "
        f"이 중 Scope 1(직접) {s1a/total_all*100:.1f}%, Scope 2(에너지) {s2a/total_all*100:.1f}%, "
        f"Scope 3(공급망) {s3a/total_all*100:.1f}%의 비중을 차지합니다.",
        f"총배출량 상위 3개 산업({', '.join(top3['산업'].tolist())})이 전체 공급망 탄소 발자국의 "
        f"{top3_share:.1f}%를 차지하여, 소수 핵심 산업에 배출이 집중되는 구조입니다.",
        f"Scope 1 비중이 가장 높은 산업은 '{df_dash.loc[s1_sh.idxmax(),'산업']}'({s1_sh.max():.0f}%)로 공정, 연소 등 직접배출 관리가 시급하며, "
        f"Scope 3 비중이 가장 높은 산업은 '{df_dash.loc[s3_sh.idxmax(),'산업']}'({s3_sh.max():.0f}%)로 공급망 관리가 탄소 저감의 핵심입니다.",
        "Scope 3 비중이 높은 산업일수록 자사 시설 개선만으로는 한계가 있으므로, 원자재 조달처 및 운송 공급망 전반에 걸친 저탄소 전환 전략이 필요합니다."
    ]

    fig_intro = plt.figure(figsize=(22, 5.0), facecolor='white')
    ax_title  = fig_intro.add_axes([0, 0.7, 1, 0.3]); ax_title.axis('off')
    ax_title.text(0.5, 0.6, f"한국 산업별 Scope 1, 2, 3 탄소 배출 종합 대시보드 ({DASH_YEAR}년, {label})",
                  fontsize=20, weight='bold', ha='center', va='center', color='#1B4332')
    ax_title.text(0.5, 0.1, f"전체 {n_sectors}개 산업 | 총배출량 내림차순 정렬 | Upstream/Downstream은 탄소유발(M행렬) 기준",
                  fontsize=11, ha='center', va='center', color='#6b7280')

    ax_ins = fig_intro.add_axes([0.02, 0.0, 0.96, 0.65]); ax_ins.axis('off')
    ax_ins.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.012",
                      facecolor='#f0fdf4', edgecolor='#16a34a', linewidth=1.3, transform=ax_ins.transAxes))
    ax_ins.text(0.02, 0.85, "종합 인사이트 - 전 산업 탄소 배출 구조 분석",
                fontsize=12, weight='bold', color='#15803d', va='top')
    iy = 0.65
    for line in insight_lines:
        block = "\n  ".join(textwrap.wrap(line, width=110))
        ax_ins.text(0.02, iy, f"- {block}", fontsize=10, color='#1f2937',
                    va='top', transform=ax_ins.transAxes, linespacing=1.5)
        iy -= 0.18

    from IPython.display import display as _display

    # ── 2. 개별 산업 카드 ────────────────────────────────────────────────────
    # 파일명에 (p_type, 연도, 순위, 산업명)을 모두 포함해 추후 이미지만 재로드 가능하도록 함
    card_index = []
    for rank, (_, row) in enumerate(df_dash.iterrows(), start=1):
        fig_card = plt.figure(figsize=(22, 7.5), facecolor='white')
        draw_sector_card(fig_card, fig_card.add_gridspec(1, 1)[0], row, rank)
        safe = "".join(c for c in row['산업'] if c.isalnum() or c in " _-")
        card_fname = f'Dashboard_Card_{DASH_YEAR}_{rank:02d}_{safe}.png'
        fig_card.savefig(out_path(card_fname, output_dir),
                         dpi=200, bbox_inches='tight', facecolor='white')
        _display(fig_card)
        plt.close(fig_card)
        card_index.append({"rank": rank, "산업": row['산업'], "file": card_fname})

    # 인트로 표지 저장 + 화면 표시
    intro_fname = f'Dashboard_00_Intro_{DASH_YEAR}.png'
    fig_intro.savefig(out_path(intro_fname, output_dir),
                      dpi=300, bbox_inches='tight', facecolor='white')
    _display(fig_intro)
    plt.close(fig_intro)

    # 인덱스 파일 저장 (연도별 산업→파일명 매핑; 위젯이 아니라 show_dashboard_year() 등에서 사용)
    import json as _json
    idx_path = out_path(f'Dashboard_Index_{p_type}_{DASH_YEAR}.json', output_dir)
    with open(idx_path, 'w', encoding='utf-8') as f:
        _json.dump({"year": DASH_YEAR, "p_type": p_type,
                    "intro_file": intro_fname, "cards": card_index},
                   f, ensure_ascii=False, indent=2)

    print("✅ 모든 산업의 대시보드 카드가 고화질로 개별 생성/저장 되었습니다.")


# ═══════════════════════════════════════════════════════════════════════════════
# 9. 경상가격 전용 — 산업별 Scope 종합 대시보드
# ═══════════════════════════════════════════════════════════════════════════════

def plot_dashboard_current(
    df_scope: pd.DataFrame,
    year: str = "2023",
    top_n: int = 5,
    output_dir: str = "./output",
):
    """
    경상가격 산업별 Scope 1/2/3 종합 대시보드 (EPA-IDA 스타일).
    '기타' 제외는 build_scope_table / plot_scope_bar_donut_by_year 단계에서 처리.
    """
    mpl.rcParams["figure.dpi"] = 200
    MY_COLORS = ["#1B4332", "#2D6A4F", "#74C69D"]
    df_dash   = df_scope.sort_values("합계", ascending=False).reset_index(drop=True)
    n_sectors = len(df_dash)

    # 이 함수는 경상가격 전용이므로 p_type='경상' 으로 고정해 복합키를 조회한다
    # (run_eeio() 가 저장하는 키 규칙 f"{year}_{p_type}" 과 동일하게 맞춤).
    _mkey = f"{year}_경상"
    mdat  = M_matrices.get(_mkey) or M_matrices.get(year, list(M_matrices.values())[-1])
    sector_names_M = list(mdat["산업명"])
    M_mat          = mdat["M"] * 1000

    def get_M_index(name):
        try:
            return sector_names_M.index(name)
        except ValueError:
            return None

    def draw_sector_card(fig, gs_row, sector_row, rank):
        sec = sector_row["산업"]
        scope1, scope2, scope3 = sector_row["Scope1"], sector_row["Scope2"], sector_row["Scope3"]
        total_123 = scope1 + scope2 + scope3 or 1e-12

        idx = get_M_index(sec)
        if idx is not None:
            col     = M_mat[:, idx].copy(); col[idx] = 0
            up_idx  = np.argsort(col)[::-1][:top_n]
            upstream_top = [(sector_names_M[i], col[i]) for i in up_idx]
            rowv    = M_mat[idx, :].copy(); rowv[idx] = 0
            dn_idx  = np.argsort(rowv)[::-1][:top_n]
            downstream_top = [(sector_names_M[i], rowv[i]) for i in dn_idx]
        else:
            upstream_top = downstream_top = []

        gs_card  = gs_row.subgridspec(1, 4, width_ratios=[1.15, 0.20, 0.9, 1.45], wspace=0.05)
        ax_left  = fig.add_subplot(gs_card[0, 0]); ax_left.axis("off")
        ax_left.set_xlim(0, 1); ax_left.set_ylim(-0.6, 1.15)

        tlen      = len(f"#{rank}  {sec}")
        tf        = 16 if tlen <= 13 else (14 if tlen <= 20 else 12.5)
        ax_left.text(0, 1.10, f"#{rank}  {sec}", fontsize=tf, weight="bold", color="#16a34a", va="top")
        ax_left.text(0, 0.99, f"한국 EEIO 분석 | {year}년 경상가격", fontsize=10, color="#6b7280", va="top")

        by = 0.88
        for xo, label_txt, val_txt, bg, ec, tc in [
            (0,    "Scope 1+2",           f"{scope1+scope2:.4f}", "#f3f4f6", "#d1d5db", "#111827"),
            (0.50, "Scope 1+2+3 (총배출)", f"{total_123:.4f}",     "#dcfce7", "#16a34a", "#15803d"),
        ]:
            ax_left.add_patch(FancyBboxPatch((xo, by-0.085), 0.46, 0.11, boxstyle="round,pad=0.012",
                                              facecolor=bg, edgecolor=ec, clip_on=False))
            ax_left.text(xo+0.23, by-0.012, label_txt, fontsize=10, ha="center", color=tc, weight="bold")
            ax_left.text(xo+0.23, by-0.065, val_txt,   fontsize=15, ha="center", color=tc, weight="bold")
        ax_left.text(0, by-0.135, "단위: t CO₂eq / 백만원 (경상가격)", fontsize=9, color="#9ca3af")

        gap = 0.085
        uy  = by - 0.23
        ax_left.text(0, uy, "Top 5 Upstream  (해당 산업 생산이 유발하는 타 산업 배출, Backward)", fontsize=11, weight="bold", color="#1B4332")
        for i, (nm, v) in enumerate(upstream_top or [("데이터 없음", 0)]):
            ax_left.text(0.02, uy-gap*(i+1), f"{i+1}. {nm}", fontsize=10, color="#374151")
            ax_left.text(0.98, uy-gap*(i+1), f"{v:.4f}",     fontsize=9.5, color="#6b7280", ha="right")

        dy = uy - gap * 6.2
        ax_left.text(0, dy, "Top 5 Downstream  (해당 산업 배출을 견인하는 타 산업, Forward)", fontsize=11, weight="bold", color="#2D6A4F")
        for i, (nm, v) in enumerate(downstream_top or [("데이터 없음", 0)]):
            ax_left.text(0.02, dy-gap*(i+1), f"{i+1}. {nm}", fontsize=10, color="#374151")
            ax_left.text(0.98, dy-gap*(i+1), f"{v:.4f}",     fontsize=9.5, color="#6b7280", ha="right")

        # 도넛
        ax_donut = fig.add_subplot(gs_card[0, 2])
        vals = [max(scope1, 0), max(scope2, 0), max(scope3, 0)]
        if sum(vals) <= 0: vals = [1, 0, 0]
        wedges = ax_donut.pie(vals, colors=MY_COLORS, startangle=90, radius=0.85,
                               wedgeprops={"width": 0.36, "edgecolor": "white", "linewidth": 2})[0]
        ax_donut.legend(wedges, ["Scope 1 (직접)", "Scope 2 (에너지)", "Scope 3 (공급망)"],
                         loc="center", bbox_to_anchor=(0.5, -0.2), frameon=False, fontsize=9)
        nf = 15 if len(sec) <= 6 else (13 if len(sec) <= 12 else 11)
        ax_donut.text(0, 0, textwrap.fill(sec, 10), ha="center", va="center",
                      fontsize=nf, weight="bold", linespacing=1.25)

        # 우측 설명
        ax_r = fig.add_subplot(gs_card[0, 3]); ax_r.axis("off")
        ax_r.set_xlim(0, 1.25); ax_r.set_ylim(-0.4, 1.15)
        rx   = 0.12
        s1p, s2p, s3p = scope1/total_123*100, scope2/total_123*100, scope3/total_123*100
        ax_r.text(rx, 1.10, "Scope 1, 2, 3 배출 구조", fontsize=14, weight="bold", color="#111827", va="top")
        ax_r.text(rx, 0.98, textwrap.fill(sec, 22),  fontsize=12, weight="bold", color="#16a34a", va="top", linespacing=1.3)
        n_lines = textwrap.fill(sec, 22).count("\n") + 1
        ty = 0.82 - 0.08 * max(n_lines-1, 0)
        for i, (sp, sc_label, desc) in enumerate([
            (s1p, "Scope 1", f"직접 연소, 공정 배출이 이 산업 배출량의 {s1p:.0f}%를 차지합니다."),
            (s2p, "Scope 2", f"구매 전력, 가스, 증기 사용에서 발생하는 간접 배출은 {s2p:.0f}%입니다."),
            (s3p, "Scope 3", f"나머지 {s3p:.0f}%는 공급망 전체(원자재, 운송 등)에서 유발됩니다."),
        ]):
            ty_i = ty - 0.30 * i
            ax_r.text(rx, ty_i,      f"Scope {i+1} ({sp:.0f}%)", fontsize=11.5, weight="bold", color=MY_COLORS[i])
            ax_r.text(rx, ty_i-0.08, desc, fontsize=10.5, color="#374151", va="top", linespacing=1.4)
        ax_r.text(rx, -0.30, "* 산업 집계 데이터 기반 근사치, 경상가격 기준", fontsize=9, color="#9ca3af", va="bottom")

    # ── 인트로 대시보드 ──────────────────────────────────────────────────────
    total_all = df_dash["합계"].sum() or 1e-12
    s1a, s2a, s3a = df_dash["Scope1"].sum(), df_dash["Scope2"].sum(), df_dash["Scope3"].sum()
    top3       = df_dash.head(3)
    top3_share = top3["합계"].sum() / total_all * 100
    s1_share   = (df_dash["Scope1"] / df_dash["합계"].replace(0, np.nan) * 100).fillna(0)
    s3_share   = (df_dash["Scope3"] / df_dash["합계"].replace(0, np.nan) * 100).fillna(0)

    insight_lines = [
        f"전체 {n_sectors}개 산업의 합산 탄소 발자국은 {total_all:.4f} t CO₂eq이며, "
        f"Scope 1 {s1a/total_all*100:.1f}%, Scope 2 {s2a/total_all*100:.1f}%, "
        f"Scope 3 {s3a/total_all*100:.1f}% 비중입니다.",
        f"상위 3개 산업({', '.join(top3['산업'].tolist())})이 전체의 {top3_share:.1f}%를 차지합니다.",
        f"Scope 1 비중이 가장 높은 산업: '{df_dash.loc[s1_share.idxmax(),'산업']}' ({s1_share.max():.0f}%)",
        f"Scope 3 비중이 가장 높은 산업: '{df_dash.loc[s3_share.idxmax(),'산업']}' ({s3_share.max():.0f}%)",
    ]

    fig_i = plt.figure(figsize=(22, 5.0), facecolor="white")
    ax_t  = fig_i.add_axes([0, 0.7, 1, 0.3]); ax_t.axis("off")
    ax_t.text(0.5, 0.6, f"한국 산업별 Scope 1, 2, 3 탄소 배출 종합 대시보드 ({year}년, 경상가격)",
              fontsize=26, weight="bold", ha="center", va="center", color="#1B4332")
    ax_t.text(0.5, 0.1, f"전체 {n_sectors}개 산업  |  총배출량 내림차순  |  Upstream/Downstream: M 행렬 기준",
              fontsize=13.5, ha="center", va="center", color="#6b7280")

    ax_ins = fig_i.add_axes([0.02, 0.0, 0.96, 0.65]); ax_ins.axis("off")
    ax_ins.add_patch(FancyBboxPatch((0,0), 1, 1, boxstyle="round,pad=0.012",
                                     facecolor="#f0fdf4", edgecolor="#16a34a", linewidth=1.3,
                                     transform=ax_ins.transAxes))
    ax_ins.text(0.02, 0.85, "종합 인사이트", fontsize=15, weight="bold", color="#15803d", va="top")
    iy = 0.65
    for line in insight_lines:
        ax_ins.text(0.02, iy, f"- {textwrap.fill(line, 110)}",
                    fontsize=12, color="#1f2937", va="top", linespacing=1.5)
        iy -= 0.18

    plt.savefig(out_path(f"Dashboard_00_Intro_{year}.png", output_dir), dpi=300,
                bbox_inches="tight", facecolor="white")
    plt.show()

    # ── 개별 산업 카드 ────────────────────────────────────────────────────────
    for rank, (_, row) in enumerate(df_dash.iterrows(), start=1):
        fig_c = plt.figure(figsize=(22, 7.5), facecolor="white")
        draw_sector_card(fig_c, fig_c.add_gridspec(1, 1)[0], row, rank)
        safe  = "".join(c for c in row["산업"] if c.isalnum() or c in " _-")
        fpath = out_path(f"Dashboard_Card_{rank:02d}_{safe}.png", output_dir)
        plt.savefig(fpath, dpi=200, bbox_inches="tight", facecolor="white")
        plt.show()

    print("✅ 모든 산업 대시보드 카드 생성 완료.")


# ═══════════════════════════════════════════════════════════════════════════════
# 10. 불변가격 전용 시각화
# ═══════════════════════════════════════════════════════════════════════════════

def plot_facet_timeseries(
    total_viz, pivot_BL, analysis_years: list,
    p_type: str = "불변",
    exclude_keyword: str | None = "기타",
    output_dir: str = "./output",
):
    """불변가격 STEP 9 격자 시계열 차트.
    '기타' 와 정확히 일치하는 산업만 제외하고 전 산업 시계열을 표시한다.
    """
    if exclude_keyword:
        sectors = [s for s in pivot_BL.index if str(s).strip() != exclude_keyword]
    else:
        sectors = list(pivot_BL.index)

    n_cols    = 4
    n_rows    = int(np.ceil(len(sectors) / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, n_rows * 3.4), facecolor="white")
    axes      = axes.flatten()

    year_to_pos = {y: i for i, y in enumerate(analysis_years)}

    for i, sec in enumerate(sectors):
        ax  = axes[i]
        sub = total_viz[total_viz["산업명"] == sec].sort_values("연도")
        x_pos = sub["연도"].map(year_to_pos)
        ax.plot(x_pos, sub["직접배출계수(B)"],     marker="s", ms=5, lw=1.8, color="#2563eb", label="직접배출(B)")
        ax.plot(x_pos, sub["공급망유발계수(B*L)"], marker="o", ms=5, lw=1.8, color="#16a34a", label="총배출(B×L)")
        ax.set_title(sec, fontsize=10, weight="bold")
        ax.set_xticks(range(len(analysis_years)))
        ax.set_xticklabels([str(y) for y in analysis_years])
        ax.tick_params(axis="x", rotation=45, labelsize=7.5)
        ax.tick_params(axis="y", labelsize=7.5)
        ax.grid(True, linestyle=":", alpha=0.4)
        if i == 0:
            ax.legend(fontsize=7.5, loc="best")
    for j in range(len(sectors), len(axes)):
        axes[j].axis("off")

    fig.suptitle(
        f"산업별 직접배출계수(B) vs 총배출계수(B×L) 시계열 비교\n"
        f"({p_type}가격, 전 산업, 단위: t CO₂eq / 백만원)",
        fontsize=15, weight="bold", y=1.02,
    )
    plt.tight_layout()
    fpath = out_path(f"Step_B_vs_BL_Timeseries_{p_type}.png", output_dir)
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"🖼️  저장 → {fpath}")


def plot_total_emission_timeseries(
    total_viz, pivot_BL, analysis_years: list,
    top_n: int = 8, output_dir: str = "./output",
):
    """불변가격 STEP 12 — 전 산업 총배출유발계수 시계열."""
    top_n_sectors = pivot_BL[2023].nlargest(top_n).index.tolist()
    fig, ax = plt.subplots(figsize=(15, 9), facecolor="white")
    for sec in pivot_BL.index:
        if sec in top_n_sectors:
            continue
        sub = total_viz[total_viz["산업명"] == sec].sort_values("연도")
        ax.plot(sub["연도"], sub["공급망유발계수(B*L)"], color="#d1d5db", lw=1, alpha=0.6, zorder=1)
    palette = sns.color_palette("tab10", len(top_n_sectors))
    for i, sec in enumerate(top_n_sectors):
        sub = total_viz[total_viz["산업명"] == sec].sort_values("연도")
        ax.plot(sub["연도"], sub["공급망유발계수(B*L)"],
                marker="o", ms=6, lw=2.3, color=palette[i], label=sec, zorder=3)
    ax.set_title("전 산업 총배출유발계수(B×L) 시계열 추이 (불변가격)", fontsize=15, weight="bold", pad=14)
    ax.set_xlabel("연도", fontsize=11, labelpad=8)
    ax.set_ylabel("공급망유발계수 (t CO₂eq / 백만원)", fontsize=11, labelpad=8)
    ax.set_xticks(analysis_years)
    ax.set_xticklabels([str(y) for y in analysis_years])
    ax.legend(title=f"상위 {top_n}개 산업 (2023년 기준)",
              fontsize=8.5, title_fontsize=9.5, loc="upper left", ncol=2, framealpha=0.9)
    ax.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    fpath = out_path("Step_TotalEmission_Timeseries_불변.png", output_dir)
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"🖼️  저장 → {fpath}")


def plot_scope_share_timeseries(
    total_viz, pivot_BL, analysis_years: list,
    top_n_scope3: int = 5, output_dir: str = "./output",
):
    """불변가격 STEP 13 — Scope 비중 시계열."""
    total_viz = total_viz.copy()
    total_viz["Scope1"] = total_viz["직접배출계수(B)"]
    yearly_scope     = total_viz.groupby("연도")[["Scope1", "Scope2", "Scope3"]].sum()
    yearly_scope_pct = yearly_scope.div(yearly_scope.sum(axis=1), axis=0) * 100
    SCOPE_COLORS     = ["#1B4332", "#2D6A4F", "#74C69D"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(17, 6.8), facecolor="white")
    ax1.stackplot(
        yearly_scope_pct.index,
        yearly_scope_pct["Scope1"], yearly_scope_pct["Scope2"], yearly_scope_pct["Scope3"],
        labels=["Scope 1 (직접)", "Scope 2 (에너지)", "Scope 3 (공급망)"],
        colors=SCOPE_COLORS, alpha=0.92,
    )
    ax1.set_title("전 산업 합산 Scope 1/2/3 비중 시계열", fontsize=13.5, weight="bold", pad=12)
    ax1.set_xlabel("연도", fontsize=11); ax1.set_ylabel("비중 (%)", fontsize=11)
    ax1.set_xticks(analysis_years); ax1.set_ylim(0, 100)
    ax1.legend(loc="upper right", fontsize=9.5, framealpha=0.9)
    ax1.grid(True, linestyle=":", alpha=0.3)

    top_s3  = pivot_BL[2023].nlargest(top_n_scope3).index.tolist()
    pal_s3  = sns.color_palette("Dark2", top_n_scope3)
    for i, sec in enumerate(top_s3):
        sub = total_viz[total_viz["산업명"] == sec].sort_values("연도").copy()
        sub["Scope3_pct"] = sub["Scope3"] / (sub["Scope1"] + sub["Scope2"] + sub["Scope3"]) * 100
        ax2.plot(sub["연도"], sub["Scope3_pct"], marker="o", ms=7, lw=2.2,
                 color=pal_s3[i], label=sec)
    ax2.set_title(f"주요 산업(상위 {top_n_scope3}개) Scope 3 비중 시계열", fontsize=13.5, weight="bold", pad=12)
    ax2.set_xlabel("연도", fontsize=11); ax2.set_ylabel("Scope 3 비중 (%)", fontsize=11)
    ax2.set_xticks(analysis_years); ax2.set_ylim(0, 100)
    ax2.legend(fontsize=8.5, loc="best", framealpha=0.9)
    ax2.grid(True, linestyle="--", alpha=0.4)

    plt.tight_layout()
    fpath = out_path("Step_Scope_Share_Timeseries_불변.png", output_dir)
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"🖼️  저장 → {fpath}")


# ═══════════════════════════════════════════════════════════════════════════════
# 11. 원-샷 초기화 함수 — 노트북에서 단 한 줄로 전체 파이프라인 실행
# ═══════════════════════════════════════════════════════════════════════════════

_STATE: dict = {}   # 모듈 수준 상태 저장소 (노트북 커널 재기동 전까지 유지)


def init_pipeline(
    ghg_file: str = "2015-2023산업별_온실가스_추정_FF_송부.xlsx",
    io_files: dict | None = None,
    analysis_years: list | None = None,
    exclude: str = "기타",
    output_dir: str = "./output",
) -> None:
    """
    전체 파이프라인을 한 번에 초기화합니다.

    노트북에서 `import eeio_core_최종 as ec; ec.init_pipeline()` 한 줄로 호출.
    결과는 모듈 내부 _STATE 에 저장되며, 이후 show_*() 함수들이 참조합니다.
    """
    if io_files is None:
        io_files = {str(y): f"{y}_투입산출표_경상_불변(최종).xlsx" for y in range(2015, 2024)}
    if analysis_years is None:
        analysis_years = list(range(2015, 2024))

    os.makedirs(output_dir, exist_ok=True)
    _STATE["output_dir"]     = output_dir
    _STATE["io_files"]       = io_files
    _STATE["analysis_years"] = analysis_years
    _STATE["exclude"]        = exclude

    print("=" * 60)
    print("  EEIO 파이프라인 초기화 시작")
    print("=" * 60)

    # ── GHG 로드 ──────────────────────────────────────────────────────────
    ghg = load_ghg_data(ghg_file)
    _STATE["ghg"] = ghg

    # ── 경상가격 ──────────────────────────────────────────────────────────
    print("\n[경상가격] 전체 연도 EEIO 계산 중...")
    fr_c = run_all_years(io_files, analysis_years, ghg, p_type="경상",
                         exclude_keyword=exclude, output_dir=output_dir)
    tv_c, pBL_c, pB_c = build_viz_data(fr_c)
    _STATE.update(final_c=fr_c, viz_c=tv_c, pBL_c=pBL_c, pB_c=pB_c)

    # ── 불변가격 ──────────────────────────────────────────────────────────
    print("\n[불변가격] 전체 연도 EEIO 계산 중...")
    fr_k = run_all_years(io_files, analysis_years, ghg, p_type="불변",
                         exclude_keyword=exclude, output_dir=output_dir)
    tv_k, pBL_k, pB_k = build_viz_data(fr_k)
    _STATE.update(final_k=fr_k, viz_k=tv_k, pBL_k=pBL_k, pB_k=pB_k)

    print("\n✅ 초기화 완료 — show_*() 함수를 호출해 결과를 확인하세요.")


# ── 상태 접근 헬퍼 ────────────────────────────────────────────────────────────
def _s(key):
    if not _STATE:
        raise RuntimeError("먼저 ec.init_pipeline() 을 실행하세요.")
    return _STATE[key]


# ═══════════════════════════════════════════════════════════════════════════════
# 12. show_*() — 노트북 셀 하나 = 시각화 하나
# ═══════════════════════════════════════════════════════════════════════════════

def show_ghg_table():
    """온실가스 데이터 미리보기 (t CO₂eq)."""
    display(HTML(
        "<div style='background-color:#f0fdf4;padding:10px;"
        "border-left:5px solid #16a34a;margin-bottom:10px'>"
        "<b>📋 온실가스 데이터 미리보기 (단위: t CO₂eq)</b></div>"
    ))
    display((_s("ghg") * 1000).round(1))


# ── 경상가격 ──────────────────────────────────────────────────────────────────

def show_c_heatmaps():
    plot_heatmaps(_s("pBL_c"), _s("pB_c"), p_type="경상", output_dir=_s("output_dir"))

def show_c_bars():
    plot_bar_charts(_s("viz_c"), _s("analysis_years"), p_type="경상", output_dir=_s("output_dir"))

def show_c_timeseries():
    plot_timeseries(_s("viz_c"), _s("pBL_c"), _s("pB_c"), _s("analysis_years"),
                    p_type="경상", output_dir=_s("output_dir"))

def show_c_pivot():
    """경상가격 전 산업 × 전 연도 피벗 결과표."""
    excl = _s("exclude")
    pBL  = _s("pBL_c"); pB = _s("pB_c")
    mask = pBL.index.str.strip() != excl
    display(HTML(
        "<div style='background-color:#f0fdf4;padding:10px;"
        "border-left:5px solid #16a34a;margin-bottom:10px'>"
        "<b>📋 공급망 유발계수 (B×L) — 전 산업 × 전 연도 (경상가격)</b><br>"
        "<span style='font-size:12px;color:#6b7280'>단위: t CO₂eq / 백만원</span></div>"
    ))
    display(pBL[mask].round(5))
    display(HTML(
        "<div style='margin-top:16px;background-color:#eff6ff;padding:10px;"
        "border-left:5px solid #2563eb;margin-bottom:10px'>"
        "<b>📋 직접 배출계수 (B) — 전 산업 × 전 연도 (경상가격)</b><br>"
        "<span style='font-size:12px;color:#6b7280'>단위: t CO₂eq / 백만원</span></div>"
    ))
    display(pB[mask].round(5))


def show_c_scope(year: str = "2023"):
    """경상가격 Scope 1/2/3 도넛+바 차트 + 표 (연도 지정)."""
    df_s = plot_scope_bar_donut_by_year(
        _s("final_c"), _s("io_files"), p_type="경상",
        year=year, exclude_keyword=_s("exclude"), output_dir=_s("output_dir"),
    )
    display(HTML(
        "<div style='background-color:#fffbeb;padding:10px;"
        f"border-left:5px solid #d97706;margin-top:16px;margin-bottom:8px'>"
        f"<b>📋 {year}년 전 산업 Scope 1/2/3 배출 구조 (경상가격)</b><br>"
        "<span style='font-size:12px;color:#6b7280'>투입가치=백만원, Scope=t CO₂eq</span></div>"
    ))
    display(df_s.set_index("산업").round(
        {"투입가치": 0, "Scope1": 4, "Scope2": 4, "Scope3": 4, "합계": 4, "탄소집약도": 6}
    ))
    return df_s


def show_c_dashboard(year: str = "2023"):
    """경상가격 EPA-IDA 스타일 대시보드 (연도 지정)."""
    df_s = plot_scope_bar_donut_by_year(
        _s("final_c"), _s("io_files"), p_type="경상",
        year=year, exclude_keyword=_s("exclude"), output_dir=_s("output_dir"),
    )
    plot_dashboard_current(df_s, year=year, output_dir=_s("output_dir"))


# ── 불변가격 ──────────────────────────────────────────────────────────────────

def show_k_pivot():
    """불변가격 전 산업 × 전 연도 피벗 결과표."""
    excl = _s("exclude")
    pBL  = _s("pBL_k"); pB = _s("pB_k")
    mask = pBL.index.str.strip() != excl
    display(HTML(
        "<div style='background-color:#f0fdf4;padding:10px;"
        "border-left:5px solid #16a34a;margin-bottom:10px'>"
        "<b>📋 공급망 유발계수 (B×L) — 전 산업 × 전 연도 (불변가격)</b><br>"
        "<span style='font-size:12px;color:#6b7280'>단위: t CO₂eq / 백만원</span></div>"
    ))
    display(pBL[mask].round(5))
    display(HTML(
        "<div style='margin-top:16px;background-color:#eff6ff;padding:10px;"
        "border-left:5px solid #2563eb;margin-bottom:10px'>"
        "<b>📋 직접 배출계수 (B) — 전 산업 × 전 연도 (불변가격)</b><br>"
        "<span style='font-size:12px;color:#6b7280'>단위: t CO₂eq / 백만원</span></div>"
    ))
    display(pB[mask].round(5))

def show_k_heatmaps():
    plot_heatmaps(_s("pBL_k"), _s("pB_k"), p_type="불변", output_dir=_s("output_dir"))

def show_k_bars():
    plot_bar_charts(_s("viz_k"), _s("analysis_years"), p_type="불변", output_dir=_s("output_dir"))

def show_k_timeseries():
    plot_timeseries(_s("viz_k"), _s("pBL_k"), _s("pB_k"), _s("analysis_years"),
                    p_type="불변", output_dir=_s("output_dir"))

def show_k_scope(year: str = "2023"):
    """불변가격 Scope 1/2/3 도넛+바 차트 + 표 (연도 지정)."""
    df_s = plot_scope_bar_donut_by_year(
        _s("final_k"), _s("io_files"), p_type="불변",
        year=year, exclude_keyword=_s("exclude"), output_dir=_s("output_dir"),
    )
    display(HTML(
        "<div style='background-color:#fffbeb;padding:10px;"
        f"border-left:5px solid #d97706;margin-top:16px;margin-bottom:8px'>"
        f"<b>📋 {year}년 전 산업 Scope 1/2/3 배출 구조 (불변가격)</b><br>"
        "<span style='font-size:12px;color:#6b7280'>투입가치=백만원, Scope=t CO₂eq</span></div>"
    ))
    display(df_s.set_index("산업").round(
        {"투입가치": 0, "Scope1": 4, "Scope2": 4, "Scope3": 4, "합계": 4, "탄소집약도": 6}
    ))
    return df_s

def show_k_facet():
    plot_facet_timeseries(_s("viz_k"), _s("pBL_k"), _s("analysis_years"), output_dir=_s("output_dir"))

def show_k_total_ts():
    plot_total_emission_timeseries(_s("viz_k"), _s("pBL_k"), _s("analysis_years"), output_dir=_s("output_dir"))

def show_k_scope_share():
    plot_scope_share_timeseries(_s("viz_k"), _s("pBL_k"), _s("analysis_years"), output_dir=_s("output_dir"))


# ═══════════════════════════════════════════════════════════════════════════════
# 11. EEIO 행렬 엑셀 저장 (예시 파일 구조 그대로)
# ═══════════════════════════════════════════════════════════════════════════════

def _find_sector_start(codes: np.ndarray) -> int:
    """코드 배열에서 'A' 섹터의 행 인덱스를 반환."""
    hits = np.where(codes == 'A')[0]
    if len(hits) == 0:
        raise ValueError("'A' 코드를 찾을 수 없습니다.")
    return int(hits[0])


def _find_output_col(df: pd.DataFrame, start_row: int) -> int:
    """총산출 열 인덱스를 동적으로 탐색."""
    keywords = ['총산출', '산출액', '공급액']
    for h_row in range(min(5, start_row)):
        for col_idx, cell in enumerate(df.iloc[h_row, :].tolist()):
            if any(kw in str(cell) for kw in keywords):
                return col_idx
    # 헤더에서 못 찾으면: 데이터 행의 마지막 NaN 아닌 열
    row_vals = df.iloc[start_row, :].tolist()
    for ci in range(len(row_vals)-1, -1, -1):
        if row_vals[ci] is not None and str(row_vals[ci]).strip() not in ('', 'nan', 'NaN'):
            return ci
    return 44


def _read_full_total_table(file_name: str, p_type: str):
    """
    총거래표 시트를 '있는 그대로' 전부 읽어 각 구성요소를 분리해 반환한다.

    실제 예시 결과 파일(2023_KR_EEIO불변__최종f.xlsx)의 T_mat_GHG 시트를
    역산해 확인한 실제 원본표 구조:

      행 구성 (위→아래):
        1) 산업 34행 (A~T, 코드+이름)
        2) 9590  중간투입계
        3) 9610  피용자보수
        4) 9621  영업잉여
        5) 9622  고정자본소모
        6) 9630  생산세(보조금공제)
        7) (조정항)  ← ★ 불변가격에만 존재, 코드가 없는(None) 행. 경상가격에는 없음.
        8) 9690  부가가치계
        9) 9790  총투입계

      열 구성 (좌→우):
        코드, 이름, A~T(34개 산업), 9090(중간수요계),
        9111~9140(최종수요 성분), 9190(최종수요계), 9290(총수요계),
        9310(총산출), 9321~9332(수입/생산물세/수입계/도소매마진/화물운임),
        9390(총공급계)

    Returns
    -------
    dict with keys:
      sector_codes, sector_names (34개 산업)
      col_codes (전체 열 코드 리스트, A부터 9390까지)
      value_rows: [(row_code_or_None, row_label, [values...]), ...]  전체 행 순서 보존
      X (9790 총투입계, 산업 열만큼의 벡터)
      has_adjustment (조정항 존재 여부)
    """
    xls = pd.ExcelFile(file_name, engine="openpyxl")
    sheet_T = _pick_sheet(xls, "총거래표", p_type)
    df_T = pd.read_excel(file_name, sheet_name=sheet_T, header=None, engine="openpyxl")

    # 코드 헤더 행 탐색: 'A' 와 '9390' 이 모두 있는 행
    code_row_idx = None
    cols_row = None
    for i in range(len(df_T)):
        str_vals = [str(v).strip() if v is not None else None for v in df_T.iloc[i].tolist()]
        if 'A' in str_vals and '9390' in str_vals:
            code_row_idx = i
            cols_row = str_vals
            break
    if code_row_idx is None:
        # 9390 이 없는 경우(수입/국산거래표 등 짧은 표) 는 지원하지 않음
        raise ValueError("총거래표에서 코드 헤더 행('A'~'9390' 포함)을 찾지 못했습니다.")

    col_start = cols_row.index('A')
    col_end   = len(cols_row) - 1
    while cols_row[col_end] is None:
        col_end -= 1
    col_codes = cols_row[col_start:col_end + 1]

    # 산업 블록(A~T) 행 범위
    all_codes = df_T.iloc[:, 0].astype(str).str.strip().values
    row_A = int(np.where(all_codes == 'A')[0][np.where(np.where(all_codes == 'A')[0] > code_row_idx)[0][0]])
    row_T = int(np.where(all_codes == 'T')[0][np.where(np.where(all_codes == 'T')[0] > code_row_idx)[0][0]])
    size = row_T - row_A + 1

    sector_codes = all_codes[row_A:row_T + 1]
    sector_names = df_T.iloc[row_A:row_T + 1, 1].values.astype(str)

    # X: 9790 총투입계 (산업 열 구간만)
    row_9790 = None
    for r in range(row_T + 1, len(df_T)):
        if str(df_T.iloc[r, 0]).strip() == '9790':
            row_9790 = r
            break
    if row_9790 is None:
        raise ValueError("총거래표에서 '9790'(총투입계) 행을 찾지 못했습니다.")
    X = df_T.iloc[row_9790, col_start:col_start + size].values.astype(float)

    # 조정항 존재 여부: 9630 다음 행이 코드 없이(None) '조정항' 이라는 라벨을 가지는지
    row_9630 = None
    for r in range(row_T + 1, row_9790 + 1):
        if str(df_T.iloc[r, 0]).strip() == '9630':
            row_9630 = r
            break
    has_adjustment = False
    if row_9630 is not None:
        next_code = df_T.iloc[row_9630 + 1, 0]
        next_label = str(df_T.iloc[row_9630 + 1, 1]).strip()
        if (pd.isna(next_code) or next_code is None) and next_label == '조정항':
            has_adjustment = True

    # 전체 행(산업 34개 + 9590 + 9610 + 9621 + 9622 + 9630 + [조정항] + 9690 + 9790)을
    # 순서 그대로 (라벨, 값배열) 형태로 수집 — T_mat_GHG 시트 그대로 재현하는 용도
    value_rows = []
    for r in range(row_A, row_9790 + 1):
        code = df_T.iloc[r, 0]
        label = df_T.iloc[r, 1]
        vals = df_T.iloc[r, col_start:col_start + len(col_codes)].values.astype(float)
        code_str = None if (pd.isna(code) or code is None) else str(code).strip()
        label_str = "" if (pd.isna(label) or label is None) else str(label).strip()
        value_rows.append((code_str, label_str, vals))

    return {
        "sector_codes": sector_codes,
        "sector_names": sector_names,
        "size": size,
        "col_codes": col_codes,
        "value_rows": value_rows,
        "X": X,
        "has_adjustment": has_adjustment,
        "row_A": row_A, "row_T": row_T,
    }


def _read_full_side_table(file_name: str, p_type: str, keyword: str, size: int):
    """
    국산거래표 / 수입거래표 시트를 '있는 그대로' 전부 읽는다.
    (9290 까지만 있는 짧은 표 — T_mat_GHG 처럼 부가가치 항목은 없음)

    Returns
    -------
    dict: sector_codes, sector_names, col_codes, matrix(2D, 산업x열),
          fd_matrix(최종수요 9111~9190 구간), fd_col_codes
    """
    xls = pd.ExcelFile(file_name, engine="openpyxl")
    try:
        sheet = _pick_sheet(xls, keyword, p_type)
    except ValueError:
        sheet = _pick_sheet(xls, keyword, p_type.replace("경상", "").replace("불변", ""))
    df = pd.read_excel(file_name, sheet_name=sheet, header=None, engine="openpyxl")

    code_row_idx = None
    cols_row = None
    for i in range(len(df)):
        str_vals = [str(v).strip() if v is not None else None for v in df.iloc[i].tolist()]
        if 'A' in str_vals and '9290' in str_vals:
            code_row_idx = i
            cols_row = str_vals
            break
    if code_row_idx is None:
        raise ValueError(f"{keyword} 시트에서 코드 헤더 행을 찾지 못했습니다.")

    col_start = cols_row.index('A')
    col_end   = cols_row.index('9290')
    col_codes = cols_row[col_start:col_end + 1]

    all_codes = df.iloc[:, 0].astype(str).str.strip().values
    a_hits = np.where(all_codes == 'A')[0]
    t_hits = np.where(all_codes == 'T')[0]
    row_A = int(a_hits[np.where(a_hits > code_row_idx)[0][0]])
    row_T = int(t_hits[np.where(t_hits > code_row_idx)[0][0]])

    sector_codes = all_codes[row_A:row_T + 1]
    sector_names = df.iloc[row_A:row_T + 1, 1].values.astype(str)

    matrix = df.iloc[row_A:row_T + 1, col_start:col_start + len(col_codes)].values.astype(float)

    col_fd_start = cols_row.index('9111')
    col_fd_end   = cols_row.index('9190')
    fd_col_codes = cols_row[col_fd_start:col_fd_end + 1]
    fd_matrix = df.iloc[row_A:row_T + 1, col_fd_start:col_fd_end + 1].values.astype(float)

    return {
        "sector_codes": sector_codes,
        "sector_names": sector_names,
        "col_codes": col_codes,
        "matrix": matrix,
        "fd_matrix": fd_matrix,
        "fd_col_codes": fd_col_codes,
    }


def build_eeio_matrices(
    year,
    file_name: str,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
) -> dict | None:
    """
    EEIO 행렬을 엑셀 저장용 DataFrame 묶음으로 구성 (예시 결과 파일과 동일 포맷).

    ★ run_eeio() 가 계산한 A/L/M/M_X/B/BL/X 를 그대로 재사용하므로
      대시보드(plot_dashboard_with_scope)와 완전히 동일한 숫자를 씁니다.

    ★ T_mat_GHG 시트는 총거래표 원본을 처음부터 끝까지(산업행 34개 +
      중간투입계 + 피용자보수/영업잉여/고정자본소모/생산세 + [조정항] +
      부가가치계 + 총투입계) 그대로 복제한  :  마지막에 GHG 행 하나만 추가합니다.
      조정항 행은 불변가격 원본에만 존재하고 경상가격에는 없으며, 이는
      원본표에 있는 그대로를 따릅니다(임의로 추가/제거하지 않음).

    ★ I_mat, D_mat 시트는 수입거래표/국산거래표 원본을 그대로(9290 까지)
      복제합니다.

    Parameters
    ----------
    year       : 분석 연도
    file_name  : 해당 연도 IO 엑셀 경로 (총거래표/국산거래표/수입거래표
                 시트를 모두 포함해야 함)
    ghg_df     : load_ghg_data() 반환값
    p_type     : '경상' 또는 '불변'

    Returns
    -------
    dict | None. 키: T_mat_GHG, Ad, Lf, M, M_X, M_Fd, Scopes, I_mat, D_mat
    """
    year_key = str(year)
    mkey = f"{year_key}_{p_type}"
    try:
        # ── 1) run_eeio() 결과 확보 (대시보드와 동일 계산 재사용) ──────────
        # ★ 키에 p_type 을 포함해 조회한다. 경상/불변을 같은 연도로 번갈아
        #   계산해도 서로 덮어쓰지 않고 각자의 캐시를 갖는다 (버그 수정:
        #   예전에는 연도만으로 캐시를 조회해서, 예를 들어 대시보드를
        #   "경상→불변" 순서로 만든 뒤 "경상" 엑셀을 저장하면 캐시에 남아있는
        #   불변가격 값을 그대로 재사용해버리는 문제가 있었다).
        if mkey not in M_matrices:
            _ = run_eeio(year_key, file_name, ghg_df, p_type=p_type, exclude_keyword=None)
            if mkey not in M_matrices:
                print(f"  ⚠️ {year}년 {p_type}: run_eeio() 계산 실패로 M 행렬을 만들 수 없습니다.")
                return None

        mdat   = M_matrices[mkey]
        A      = mdat["A"]
        L      = mdat["L"]
        M      = mdat["M"]
        M_X    = mdat["M_X"]
        B      = mdat["B"]
        BL     = mdat["BL"]
        X      = mdat["X"]
        sector_codes = mdat["코드"]
        sector_names = mdat["산업명"]
        labels       = mdat["라벨"]
        size         = len(sector_codes)

        idx_D_arr = np.where(sector_codes == "D")[0]
        D_rel     = int(idx_D_arr[0]) if len(idx_D_arr) else None

        # ── Scope 1/2/3 (run_eeio 와 동일 정의) ─────────────────────────────
        scope_1 = B.copy()
        if D_rel is not None:
            scope_2 = M[D_rel, :].copy()
            scope_2[D_rel] -= scope_1[D_rel]
            scope_2 = np.maximum(scope_2, 0.0)
        else:
            scope_2 = np.zeros(size)
        scope_3 = np.maximum(BL - scope_1 - scope_2, 0.0)

        # ── 2) 총거래표 원본 통째로 읽기 (T_mat_GHG 용) ─────────────────────
        # 열 라벨: 산업 코드는 (코드, 이름) MultiIndex, 9090/9111 등 특수 코드는
        # (코드, "") 형태로 통일해 전체 컬럼을 하나의 MultiIndex로 만든다.
        tot_info = _read_full_total_table(file_name, p_type)
        col_name_map = dict(zip(sector_codes.tolist(), sector_names.tolist()))

        def _col_multi(code_list):
            return pd.MultiIndex.from_tuples(
                [(c, col_name_map.get(c, "")) for c in code_list],
                names=["코드", "산업명"],
            )

        # run_eeio() 에서 얻은 산업 순서와 총거래표 산업 순서가 같은지 검증
        if not np.array_equal(tot_info["sector_codes"], sector_codes):
            print(f"  ⚠️ {year}년 {p_type}: 총거래표 산업코드 순서가 run_eeio() 결과와 다릅니다. "
                  f"T_mat_GHG 시트를 건너뜁니다.")
            df_T_out = pd.DataFrame(index=labels)
        else:
            col_codes_full = tot_info["col_codes"]
            col_labels_full = _col_multi(col_codes_full)
            ghg_row_vals = np.full(len(col_codes_full), np.nan)
            ghg_row_vals[:size] = mdat.get("GHG_raw", B * X)  # t 단위 배출량 (Scope1과 동일)

            # 행 라벨: 산업행은 (코드, 이름), 9590/9610 등은 (코드, 라벨),
            # 조정항처럼 코드가 없는 행은 ("", 라벨)
            row_tuples = []
            row_vals_list = []
            for code, label, vals in tot_info["value_rows"]:
                if code in sector_codes.tolist():
                    idx = sector_codes.tolist().index(code)
                    row_key = (sector_codes[idx], sector_names[idx])
                elif code is None:
                    row_key = ("", label)
                else:
                    row_key = (code, label)
                row_tuples.append(row_key)
                row_vals_list.append(vals)

            row_index_full = pd.MultiIndex.from_tuples(row_tuples, names=["코드", "산업명"])
            df_T_out = pd.DataFrame(row_vals_list, index=row_index_full, columns=col_labels_full)
            ghg_row_key = pd.MultiIndex.from_tuples([("GHG", "탄소배출량(tco2eq.)")], names=["코드", "산업명"])
            df_ghg_row = pd.DataFrame([ghg_row_vals], index=ghg_row_key, columns=col_labels_full)
            df_T_out = pd.concat([df_T_out, df_ghg_row])

        # ── 3) 국산거래표 / 수입거래표 원본 통째로 읽기 (D_mat / I_mat 용) ──
        col_sum_key   = ("9590", "중간투입계")

        def _append_colsum_row(df, key=col_sum_key):
            """DataFrame 맨 아래에 열 합계(중간투입계) 행을 하나 추가."""
            if df.empty:
                return df
            colsum = df.sum(axis=0)
            colsum_df = colsum.to_frame().T
            colsum_df.index = pd.MultiIndex.from_tuples([key], names=["코드", "산업명"])
            return pd.concat([df, colsum_df])

        try:
            dom_info = _read_full_side_table(file_name, p_type, "국산거래표", size)
            df_D_out = pd.DataFrame(
                dom_info["matrix"], index=labels, columns=_col_multi(dom_info["col_codes"])
            )
            df_D_out = _append_colsum_row(df_D_out)
            F_d = dom_info["fd_matrix"]
            fd_col_codes = dom_info["fd_col_codes"]
        except (ValueError, IndexError) as e:
            print(f"  ⚠️ {year}년 {p_type}: 국산거래표 파싱 실패 ({e})")
            df_D_out = pd.DataFrame(index=labels)
            F_d = None
            fd_col_codes = None

        try:
            imp_info = _read_full_side_table(file_name, p_type, "수입거래표", size)
            df_I_out = pd.DataFrame(
                imp_info["matrix"], index=labels, columns=_col_multi(imp_info["col_codes"])
            )
            df_I_out = _append_colsum_row(df_I_out)
        except (ValueError, IndexError) as e:
            print(f"  ⚠️ {year}년 {p_type}: 수입거래표 파싱 실패 ({e})")
            df_I_out = pd.DataFrame(index=labels)

        # ── 4) M_Fd (최종수요 성분별 배출량, run_eeio 의 M 기반) ────────────
        # 최종수요 열(9111~9190)은 산업 코드가 아니므로 (코드, "") 형태로 표시
        if F_d is not None:
            M_Fd = M @ F_d
            fd_col_multi = pd.MultiIndex.from_tuples(
                [(c, "") for c in fd_col_codes], names=["코드", "산업명"]
            )
            df_M_Fd = pd.DataFrame(M_Fd, index=labels, columns=fd_col_multi)
            df_M_Fd = _append_colsum_row(df_M_Fd)
        else:
            df_M_Fd = pd.DataFrame(index=labels)

        # ── 5) DataFrame 생성 (행: 코드행+이름행 / 열: 코드열+이름열 분리) ──
        # 예시 결과 파일(2023_KR_EEIO불변_최종f.xlsx) 구조 재현:
        #   - Ad, M, M.X(diag): 오른쪽 끝에 행합계(9090) 열, 맨 아래에
        #     "9590_중간투입계" 열합계 행 추가. Ad 시트에는 그 아래 한 줄 더
        #     "GHG_배출계수(tco2eq./백만원)" 행(=B 벡터)을 추가.
        #   - Lf: 레온티에프 역행렬 + 맨 아래 "9590_중간투입계"(열합계) 행.
        #     (예시 파일에 있던 diag(B)/M 중복 블록은 재현하지 않음 — 작업
        #     과정의 실수로 판단되어 제외.)
        row_sum_key   = ("9090", "행합계")
        ad_ghg_key    = ("GHG", "배출계수(tco2eq./백만원)")

        def _add_row_col_sums(df_mat, mat_np, extra_rows=()):
            """정사각 행렬 DataFrame(labels x labels)에 행합계 열, 열합계 행(들)을 추가."""
            df = df_mat.copy()
            df[row_sum_key] = df.sum(axis=1)
            extra_frames = [df]
            for row_key, row_vals in extra_rows:
                s = pd.Series(row_vals, index=labels)
                s[row_sum_key] = np.nan if np.isnan(row_vals).all() else np.nansum(row_vals)
                extra_frames.append(s.to_frame().T)
                extra_frames[-1].index = pd.MultiIndex.from_tuples([row_key], names=["코드", "산업명"])
            return pd.concat(extra_frames)

        df_Ad = pd.DataFrame(A, index=labels, columns=labels)
        df_M  = pd.DataFrame(M, index=labels, columns=labels)
        df_M_X = pd.DataFrame(M_X, index=labels, columns=labels)

        df_Ad = _add_row_col_sums(
            df_Ad, A,
            extra_rows=[(col_sum_key, A.sum(axis=0)), (ad_ghg_key, B)],
        )
        df_M = _add_row_col_sums(
            df_M, M,
            extra_rows=[(col_sum_key, M.sum(axis=0))],
        )
        df_M_X = _add_row_col_sums(
            df_M_X, M_X,
            extra_rows=[(col_sum_key, M_X.sum(axis=0))],
        )

        df_Lf = pd.DataFrame(L, index=labels, columns=labels)
        df_Lf[row_sum_key] = df_Lf.sum(axis=1)
        df_Lf = _append_colsum_row(df_Lf)

        # ── Scopes: 예시 결과 파일과 동일 포맷 (절대량 t CO2eq) ─────────────
        # 컬럼: scope1, scope2, scope3, 총배출 (코드/산업명은 인덱스 2단으로 분리)
        # (원단위 B/M 이 아니라 X를 곱한 절대 배출량 기준 — 예시 파일 Scopes
        #  시트의 scope1 값이 GHG 원본(t)과 정확히 같은 것으로 확인됨)
        scope_1_abs = scope_1 * X
        scope_2_abs = scope_2 * X
        scope_3_abs = scope_3 * X
        total_abs   = scope_1_abs + scope_2_abs + scope_3_abs

        df_Scopes = pd.DataFrame(
            {
                'scope1': scope_1_abs,
                'scope2': scope_2_abs,
                'scope3': scope_3_abs,
                '총배출': total_abs,
            },
            index=labels,
        )
        # 소계(맨 아래 합계) 행 추가 — 예시 파일 40행 '소계' 그대로 재현
        subtotal_key = pd.MultiIndex.from_tuples([("", "소계")], names=["코드", "산업명"])
        subtotal = pd.DataFrame(
            {
                'scope1': [scope_1_abs.sum()],
                'scope2': [scope_2_abs.sum()],
                'scope3': [scope_3_abs.sum()],
                '총배출': [total_abs.sum()],
            },
            index=subtotal_key,
        )
        df_Scopes = pd.concat([df_Scopes, subtotal])

        return {
            "T_mat_GHG": df_T_out,
            "Ad":        df_Ad,
            "Lf":        df_Lf,
            "M":         df_M,
            "M_X":       df_M_X,
            "M_Fd":      df_M_Fd,
            "Scopes":    df_Scopes,
            "I_mat":     df_I_out,
            "D_mat":     df_D_out,
        }

    except Exception as exc:
        import traceback
        print(f"❌ {year}년 {p_type} 행렬 계산 오류: {exc}")
        traceback.print_exc()
        return None


def save_eeio_excel(
    year,
    file_name: str,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
    output_dir: str = "./output",
) -> str | None:
    """
    단일 연도 EEIO 결과를 엑셀로 저장 (예시 파일 2023_KR_EEIO불변_최종f.xlsx 와 동일 포맷).

    시트 순서: 설명 / T_mat_GHG / Ad / Lf / M / M.Fd / M.X(diag) / Scopes / I_mat / D_mat

    ★ T_mat_GHG 시트는 총거래표 원본을 처음부터 끝까지(산업 34행 + 중간투입계 +
      피용자보수/영업잉여/고정자본소모/생산세 + [조정항, 불변만 존재] +
      부가가치계 + 총투입계) 그대로 복제하고 맨 아래에 GHG 배출량 행 하나만
      추가합니다. 경상가격 원본에는 조정항이 없으므로 경상 결과 파일에도
      조정항 행이 나타나지 않습니다 — 원본 그대로를 따른 것이며 임의로
      추가/삭제하지 않습니다.
    ★ M.X(diag) 시트가 금액 유발행렬(M x diag(X), 단위 t CO2eq)입니다.
      대시보드 카드의 Upstream/Downstream Top 5 값과 정확히 대조 가능합니다.
    ★ 모든 시트의 행/열이 "산업코드_산업명" (예: D_전력, 가스 및 증기) 형태로
      표기되어 코드와 이름을 동시에 확인할 수 있습니다.
    ★ 대시보드(plot_dashboard_with_scope)와 완전히 동일한 A/L/M 계산 결과를
      사용하므로(run_eeio() 재사용), Upstream/Downstream 이 대시보드와
      엑셀에서 일치합니다.

    Returns
    -------
    저장된 파일 경로 (str) 또는 None (오류 시)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mats = build_eeio_matrices(year, file_name, ghg_df, p_type)
    if mats is None:
        return None

    label = "경상" if p_type == "경상" else "불변"
    fpath = out_path(f"{year}_KR_EEIO_{label}.xlsx", output_dir)

    # ── 스타일 헬퍼 ────────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="1B4332")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    IDX_FILL = PatternFill("solid", fgColor="D8F3DC")
    IDX_FONT = Font(bold=True, color="1B4332", size=10)
    BODY_FONT = Font(size=9, name="Arial")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_matrix(ws, df: pd.DataFrame, title: str, note: str = ""):
        """
        DataFrame을 시트에 작성. 인덱스/컬럼이 (코드, 이름) MultiIndex인 경우
        행은 [코드열, 이름열] 2개 열로, 열은 [코드행, 이름행] 2개 행으로
        분리하여 표시한다 (예시 파일의 5행=코드, 6행=이름 구조와 동일).
        MultiIndex 가 아닌 단순 인덱스/컬럼(Scopes 등)은 1열/1행으로 표시.
        """
        is_multi_idx = isinstance(df.index, pd.MultiIndex)
        is_multi_col = isinstance(df.columns, pd.MultiIndex)
        n_idx_cols = 2 if is_multi_idx else 1   # 행 라벨이 차지하는 열 수
        n_hdr_rows = 2 if is_multi_col else 1   # 열 라벨이 차지하는 행 수

        ws.freeze_panes = get_column_letter(n_idx_cols + 1) + str(n_hdr_rows + 2)

        # 제목행
        ws.row_dimensions[1].height = 20
        tc = ws.cell(1, 1, title)
        tc.font = Font(bold=True, size=12, color="1B4332")
        if note:
            ws.cell(1, 2, note).font = Font(size=9, color="6B7280", italic=True)

        row_offset = 2  # 헤더가 시작되는 행

        def _disp(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            return str(v)

        # ── 인덱스(행 라벨) 헤더 셀: "코드"/"산업명" 라벨 ───────────────────
        if is_multi_idx:
            ws.cell(row_offset, 1, "코드").font = HDR_FONT
            ws.cell(row_offset, 1).fill = HDR_FILL
            ws.cell(row_offset, 1).alignment = CENTER
            ws.cell(row_offset, 2, "산업명").font = HDR_FONT
            ws.cell(row_offset, 2).fill = HDR_FILL
            ws.cell(row_offset, 2).alignment = CENTER
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 26
        else:
            ws.cell(row_offset, 1, "코드").font = HDR_FONT
            ws.cell(row_offset, 1).fill = HDR_FILL
            ws.cell(row_offset, 1).alignment = CENTER
            ws.column_dimensions["A"].width = 14

        # ── 열(컬럼) 헤더: 코드행 + 이름행 2줄로 분리, 또는 단순 1줄 ─────────
        for ci, col in enumerate(df.columns, start=n_idx_cols + 1):
            if is_multi_col:
                code_v, name_v = col
                c1 = ws.cell(row_offset, ci, _disp(code_v))
                c1.font = HDR_FONT; c1.fill = HDR_FILL; c1.alignment = CENTER
                c2 = ws.cell(row_offset + 1, ci, _disp(name_v))
                c2.font = HDR_FONT; c2.fill = HDR_FILL; c2.alignment = CENTER
            else:
                c1 = ws.cell(row_offset, ci, _disp(col))
                c1.font = HDR_FONT; c1.fill = HDR_FILL; c1.alignment = CENTER
            ws.column_dimensions[get_column_letter(ci)].width = 20

        # ── 데이터행 ─────────────────────────────────────────────────────
        data_start_row = row_offset + n_hdr_rows
        for ri, (idx, row_data) in enumerate(df.iterrows(), start=data_start_row):
            if is_multi_idx:
                code_v, name_v = idx
                ic1 = ws.cell(ri, 1, _disp(code_v))
                ic1.font = IDX_FONT; ic1.fill = IDX_FILL; ic1.alignment = CENTER
                ic1.border = BORDER
                ic2 = ws.cell(ri, 2, _disp(name_v))
                ic2.font = IDX_FONT; ic2.fill = IDX_FILL; ic2.alignment = CENTER
                ic2.border = BORDER
            else:
                ic1 = ws.cell(ri, 1, _disp(idx))
                ic1.font = IDX_FONT; ic1.fill = IDX_FILL; ic1.alignment = CENTER
                ic1.border = BORDER

            for ci, val in enumerate(row_data, start=n_idx_cols + 1):
                vc = ws.cell(ri, ci)
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    vc.value = ""
                elif isinstance(val, (int, np.integer)):
                    vc.value = int(val)
                elif isinstance(val, (float, np.floating)):
                    vc.value = round(float(val), 10)
                    vc.number_format = "0.0000000000"
                else:
                    vc.value = val
                vc.font = BODY_FONT
                vc.border = BORDER

    wb = openpyxl.Workbook()

    # ── 설명 시트 ──────────────────────────────────────────────────────────
    ws_desc = wb.active
    ws_desc.title = "설명"
    desc_rows = [
        ("기호", "설명"),
        ("T_mat_GHG",   f"총거래표(산업별 온실가스 포함) — 원본표 그대로 + GHG 행"),
        ("Ad",          "국산투입계수(온실가스 배출계수 포함) — 열정규화: Z_국산거래표 / X_총투입계"),
        ("Lf",          "생산유발계수 Lf = (I-Ad)^-1"),
        ("M",           "온실가스 배출유발계수(배출계수 x Lf), 원단위 t CO2eq/백만원"),
        ("M.Fd",        "M * 최종수요벡터 (9111~9190), 단위 t CO2eq"),
        ("M.X(diag)",   "M * 생산액대각행렬 (M x diag(X)), 절대량 t CO2eq — 대시보드 Upstream/Downstream과 동일 값"),
        ("Scopes",      "Scope1,2,3 분리 (전력 이중계산 보정), 절대량 t CO2eq"),
        ("I_mat",       "수입거래표 (원본표 그대로)"),
        ("D_mat",       "국산거래표 (원본표 그대로)"),
    ]
    ws_desc.column_dimensions["A"].width = 16
    ws_desc.column_dimensions["B"].width = 70
    for ri, (k, v) in enumerate(desc_rows, 1):
        ca = ws_desc.cell(ri, 1, k)
        cb = ws_desc.cell(ri, 2, v)
        if ri == 1:
            ca.font = cb.font = Font(bold=True, size=11)
        else:
            ca.font = Font(bold=True, color="1B4332", size=10)
            cb.font = Font(size=10)

    # ── 행렬 시트들 (예시 파일과 동일 순서·시트명) ──────────────────────────
    sheet_defs = [
        ("T_mat_GHG",   mats["T_mat_GHG"],
         f"총거래표(산업별 온실가스 포함) ({year}년, {label}가격)"),
        ("Ad",          mats["Ad"],
         f"국산투입계수(온실가스 배출계수 포함) ({year}년, {label}가격)"),
        ("Lf",          mats["Lf"],
         f"생산유발계수 ({year}년, {label}가격)"),
        ("M",           mats["M"],
         f"온실가스 배출유발계수(배출계수×Lf) ({year}년, {label}가격)"),
        ("M.Fd",        mats["M_Fd"],
         f"M×최종수요벡터 ({year}년, {label}가격)"),
        ("M.X(diag)",   mats["M_X"],
         f"M×생산액대각행렬 ({year}년, {label}가격)"),
        ("Scopes",      mats["Scopes"],
         f"Scope 1/2/3 분리 ({year}년, {label}가격)"),
        ("I_mat",       mats["I_mat"],
         f"수입거래표 ({year}년, {label}가격)"),
        ("D_mat",       mats["D_mat"],
         f"국산거래표 ({year}년, {label}가격)"),
    ]
    for sname, df, title in sheet_defs:
        ws = wb.create_sheet(sname)
        _write_matrix(ws, df, title)

    wb.save(fpath)
    print(f"✅ {year}년 {label}가격 EEIO 엑셀 저장 완료 (예시 파일과 동일 포맷) → {fpath}")
    return fpath


def save_eeio_excel_all_years(
    io_files: dict,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
    output_dir: str = "./output",
) -> list[str]:
    """
    전체 연도 EEIO 결과를 연도별 엑셀 파일로 저장.

    Returns
    -------
    저장된 파일 경로 리스트
    """
    saved = []
    for year, fpath in io_files.items():
        result = save_eeio_excel(year, fpath, ghg_df, p_type, output_dir)
        if result:
            saved.append(result)
    print(f"\n✅ 총 {len(saved)}개 파일 저장 완료")
    return saved


# ═══════════════════════════════════════════════════════════════════════════════
# 11-A. ★★★ [NEW] 점검(검증) 함수 — 대시보드 ↔ 엑셀 정합성 자동 확인 ★★★
# ═══════════════════════════════════════════════════════════════════════════════
#
#  대시보드 카드의 Upstream/Downstream Top 5 값과 엑셀 M_X 시트 값이
#  실제로 일치하는지, Scope1+2+3 합이 M_X 행 합계와 일치하는지,
#  M_X 의 열 합계가 BL*X(공급망유발계수 x 총투입액)과 일치하는지를
#  자동으로 검증합니다. 노트북에서 연도만 지정해 실행하면 됩니다.
#
#  사용 예시:
#      import eeio_core_최종 as ec
#      ec.verify_eeio_consistency('2023', IO_FILES['2023'], ghg_df, p_type='경상')
# ═══════════════════════════════════════════════════════════════════════════════

def verify_eeio_consistency(
    year,
    file_name: str,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
    exclude_keyword: str | None = "기타",
    top_n: int = 3,
    tol: float = 1e-6,
) -> dict:
    """
    대시보드(run_eeio/plot_dashboard_with_scope)와 엑셀(build_eeio_matrices/
    save_eeio_excel)이 정확히 같은 숫자를 쓰는지 자동 점검합니다.

    점검 항목
    ---------
    1) M_X 행렬이 실제로 존재하고 대칭적으로 잘 구성되었는가
       (대각성분 포함, 크기 = 산업 수 x 산업 수)
    2) Scope1 + Scope2 + Scope3 == M_X 의 "해당 산업 열(column) 합계"
       (그 산업이 1백만원어치 생산할 때 유발되는 총배출 = 그 열의 합)
    3) 임의로 뽑은 top_n 개 산업에 대해, plot_dashboard_with_scope() 카드가
       그리는 것과 동일한 방식(M_X 열/행에서 argsort)으로 계산한
       Upstream/Downstream Top 5 가 재현되는지 직접 재계산해 비교
    4) M 원단위 계수와 M_X 절대량의 관계 M_X = M @ diag(X) 가 수치적으로
       성립하는지 재검증

    문제가 없으면 "✅ 통과", 문제가 있으면 "❌ 불일치" 메시지와 함께
    구체적인 산업명·수치를 출력합니다.

    Returns
    -------
    dict : {"ok": bool, "issues": [str, ...], "detail": {...}}
    """
    issues: list[str] = []
    year_key = str(year)
    mkey = f"{year_key}_{p_type}"

    print("━" * 70)
    print(f"  🔍 EEIO 정합성 점검 — {year}년 {p_type}가격")
    print("━" * 70)

    # ── 1) run_eeio() 를 실행해 M_matrices 확보 (대시보드와 동일 경로) ────
    res = run_eeio(year_key, file_name, ghg_df, p_type=p_type, exclude_keyword=exclude_keyword)
    if res is None or mkey not in M_matrices:
        issues.append("run_eeio() 계산 실패 — M 행렬을 만들 수 없습니다.")
        print("❌ run_eeio() 계산 실패. 점검을 중단합니다.")
        return {"ok": False, "issues": issues, "detail": {}}

    mdat   = M_matrices[mkey]
    M      = mdat["M"]
    M_X    = mdat["M_X"]
    B      = mdat["B"]
    BL     = mdat["BL"]
    X      = mdat["X"]
    codes  = mdat["코드"]
    names  = mdat["산업명"]
    labels = mdat["라벨"]   # MultiIndex (코드, 이름)
    size   = len(codes)

    def _lbl(i):
        """MultiIndex 라벨을 '코드_이름' 형태의 읽기 쉬운 문자열로 변환 (표시 전용)."""
        c, n = labels[i]
        return f"{c}_{n}"

    idx_D_arr = np.where(codes == "D")[0]
    D_rel = int(idx_D_arr[0]) if len(idx_D_arr) else None

    # ── 점검 1: M_X 형태 ────────────────────────────────────────────────
    if M_X.shape != (size, size):
        issues.append(f"M_X 크기 불일치: {M_X.shape} (기대: {(size, size)})")
    else:
        print(f"✅ [1] M_X 크기 정상: {M_X.shape}")

    # ── 점검 2: Scope1+2+3 == M_X 열 합계 ──────────────────────────────
    scope_1 = B.copy()
    if D_rel is not None:
        scope_2 = M[D_rel, :].copy()
        scope_2[D_rel] -= scope_1[D_rel]
        scope_2 = np.maximum(scope_2, 0.0)
    else:
        scope_2 = np.zeros(size)
    scope_3 = np.maximum(BL - scope_1 - scope_2, 0.0)
    scope_total_per_unit = scope_1 + scope_2 + scope_3          # 원단위 (1백만원당)

    X_safe = np.nan_to_num(X, nan=0.0)
    scope_total_abs = scope_total_per_unit * X_safe             # 절대량 (t CO2eq)
    M_X_col_sum      = M_X.sum(axis=0)                          # M_X 열 합계 (절대량)

    diff = np.abs(scope_total_abs - M_X_col_sum)
    max_diff_idx = int(np.argmax(diff))
    max_diff     = float(diff[max_diff_idx])

    if max_diff > tol * max(1.0, abs(scope_total_abs[max_diff_idx])):
        issues.append(
            f"Scope1+2+3(절대량) ≠ M_X 열합계: 최대 오차 {max_diff:.6g} "
            f"@ {_lbl(max_diff_idx)} "
            f"(Scope합={scope_total_abs[max_diff_idx]:.6g}, M_X열합={M_X_col_sum[max_diff_idx]:.6g})"
        )
        print(f"❌ [2] Scope 합계 vs M_X 열합계 불일치 — 최대 오차 {max_diff:.6g} @ {_lbl(max_diff_idx)}")
    else:
        print(f"✅ [2] Scope1+2+3(절대량) == M_X 열합계 (최대 오차 {max_diff:.3g}, 허용치 이내)")

    # ── 점검 3: Upstream/Downstream Top-N 재현 확인 ────────────────────
    rng = np.random.default_rng(42)
    sample_idx = rng.choice(size, size=min(top_n, size), replace=False)

    detail_samples = []
    for si in sample_idx:
        col      = M_X[:, si].copy(); col[si] = 0
        up_order = np.argsort(col)[::-1][:5]
        rowv     = M_X[si, :].copy(); rowv[si] = 0
        dn_order = np.argsort(rowv)[::-1][:5]

        up_list = [(_lbl(i), float(col[i])) for i in up_order]
        dn_list = [(_lbl(i), float(rowv[i])) for i in dn_order]

        detail_samples.append({
            "산업": _lbl(si),
            "upstream_top5": up_list,
            "downstream_top5": dn_list,
        })

        print(f"\n  📌 샘플 산업: {_lbl(si)}")
        print(f"     Upstream Top5 (M_X 열 기준, 이 산업 생산으로 유발된 타산업 배출):")
        for nm, v in up_list:
            print(f"        - {nm}: {v:,.4f} t CO2eq")
        print(f"     Downstream Top5 (M_X 행 기준, 이 산업 배출이 유발한 타산업 생산):")
        for nm, v in dn_list:
            print(f"        - {nm}: {v:,.4f} t CO2eq")

    print(f"\n✅ [3] Upstream/Downstream 재계산 완료 — 위 값이 대시보드 카드 및 "
          f"엑셀 M_X 시트와 동일해야 합니다. (동일 M_matrices[{year_key}]['M_X'] 사용)")

    # ── 점검 4: M_X == M @ diag(X) 수치 재검증 ─────────────────────────
    M_X_recompute = M @ np.diag(X_safe)
    mx_diff = np.abs(M_X - M_X_recompute).max()
    if mx_diff > tol:
        issues.append(f"M_X != M @ diag(X) — 최대 오차 {mx_diff:.6g}")
        print(f"❌ [4] M_X = M @ diag(X) 재검증 실패 — 최대 오차 {mx_diff:.6g}")
    else:
        print(f"✅ [4] M_X = M @ diag(X) 수치 재검증 통과 (최대 오차 {mx_diff:.3g})")

    print("━" * 70)
    if issues:
        print(f"❌ 점검 결과: {len(issues)}건의 불일치 발견")
        for iss in issues:
            print(f"   - {iss}")
    else:
        print("✅ 점검 결과: 모든 항목 통과 — 대시보드와 엑셀이 동일한 값을 사용합니다.")
    print("━" * 70)

    return {
        "ok": len(issues) == 0,
        "issues": issues,
        "detail": {
            "scope_total_abs": scope_total_abs,
            "M_X_col_sum": M_X_col_sum,
            "samples": detail_samples,
            "labels": labels.tolist(),
        },
    }


def verify_eeio_consistency_all_years(
    io_files: dict,
    ghg_df: pd.DataFrame,
    p_type: str = "경상",
    exclude_keyword: str | None = "기타",
) -> dict:
    """
    io_files 에 있는 모든 연도에 대해 verify_eeio_consistency() 를 실행하고
    연도별 통과 여부를 요약해 반환합니다.
    """
    results = {}
    for yr in io_files.keys():
        r = verify_eeio_consistency(
            yr, io_files[yr], ghg_df, p_type=p_type,
            exclude_keyword=exclude_keyword, top_n=1,
        )
        results[yr] = r["ok"]

    print("\n" + "═" * 70)
    print("  📋 전체 연도 점검 요약")
    print("═" * 70)
    for yr, ok in results.items():
        mark = "✅ 통과" if ok else "❌ 불일치"
        print(f"   {yr}년 : {mark}")
    print("═" * 70)
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 12. 저장된 이미지 재로드 (재계산 없이 즉시 표시)
# ═══════════════════════════════════════════════════════════════════════════════
#
#  plot_dashboard_with_scope() 등으로 한 번 생성·저장된 PNG/JSON 인덱스를
#  그대로 불러와 표시하는 함수들. 매번 IO 파일을 다시 읽고 행렬을 재계산하는
#  대신, 디스크에 저장된 이미지를 즉시 띄워서 속도를 크게 단축한다.
#
#  전제 조건: plot_dashboard_with_scope() 를 최소 1회 실행하여
#             Dashboard_Index_{p_type}_{year}.json 과 카드 PNG들이
#             output_dir 에 이미 저장되어 있어야 함.


def load_dashboard_index(p_type: str, year: str, output_dir: str = "./output") -> dict | None:
    """
    Dashboard_Index_{p_type}_{year}.json 을 읽어 {year, p_type, intro_file, cards} 반환.
    cards 는 [{"rank": int, "산업": str, "file": str}, ...] 형태.
    """
    import json
    idx_path = out_path(f"Dashboard_Index_{p_type}_{year}.json", output_dir)
    if not os.path.exists(idx_path):
        print(f"⚠️ 인덱스 파일 없음: {idx_path}\n"
              f"   → plot_dashboard_with_scope(..., p_type='{p_type}', year='{year}') 를 먼저 1회 실행하세요.")
        return None
    with open(idx_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ═══════════════════════════════════════════════════════════════════════════════
# 12-A. ★★★ [NEW] 위젯 없이 "연도 지정 → 전체 카드 순차 출력" 함수 ★★★
# ═══════════════════════════════════════════════════════════════════════════════
#
#  요청 사항: 대시보드를 위젯으로 클릭해서 보는 대신, 노트북 셀에 연도를
#  숫자/문자열로 직접 입력해서 실행하면 해당 연도의 인트로 + 전 산업 카드가
#  순서대로(rank 순) 전부 출력되어야 함. ipywidgets 의 Dropdown 상태에 의존
#  하지 않고, 함수 인자로 받은 연도를 그대로 사용해 파일을 찾아 표시하므로
#  커널을 재시작해도 "셀 재실행"만으로 동일하게 재현됩니다.
#
#  사용 예시 (노트북 셀):
#      import eeio_core_최종 as ec
#
#      # 1) 저장된 이미지가 있는 연도 → 재계산 없이 즉시 전체 출력
#      ec.show_dashboard_year(2023)                  # 경상가격, 기본값
#      ec.show_dashboard_year(2023, p_type="불변")    # 불변가격
#
#      # 2) 아직 한 번도 그리지 않은 연도라면 → 계산하면서 전체 출력
#      #    (final_c / io_files 는 init_pipeline() 실행 후 결과이거나
#      #     직접 run_all_years() 로 만든 dict 를 넣으면 됩니다)
#      ec.show_dashboard_year_recompute(2023, final_c, io_files,
#                                        p_type="경상", exclude_keyword="기타")
#
#      # 3) init_pipeline() 을 이미 실행한 경우 — 가장 간단, 알아서 처리
#      ec.show_dashboard_year_state(2023)             # 경상가격
#      ec.show_dashboard_year_state(2023, p_type="불변")
#
#      # 4) JSON 인덱스 없이 파일명 스캔만으로 표시하고 싶다면
#      ec.show_dashboard_year_v2(2023)
# ═══════════════════════════════════════════════════════════════════════════════

def show_dashboard_year(
    year,
    p_type: str = "경상",
    output_dir: str = "./output",
) -> None:
    """
    ★ [연도만 입력하면 전체 출력] ★
    저장된 이미지(JSON 인덱스 + PNG)를 기반으로 해당 연도의 대시보드
    인트로 + 전 산업 카드를 순서대로(rank 오름차순) 전부 표시합니다.

    위젯을 전혀 사용하지 않으므로, 노트북을 껐다가 켜도 이 함수를
    다시 호출하는 것만으로 동일한 결과가 재현됩니다.

    사전 조건: 해당 (p_type, year) 조합에 대해 plot_dashboard_with_scope() 가
               이전에 최소 1회 실행되어 output_dir 에
               Dashboard_Index_{p_type}_{year}.json 및 카드 PNG들이 저장되어
               있어야 합니다. 없으면 안내 메시지를 출력하고
               show_dashboard_year_recompute() 사용을 권장합니다.

    Parameters
    ----------
    year       : 연도 (int 또는 str, 예: 2023)
    p_type     : '경상' 또는 '불변'
    output_dir : 이미지/인덱스가 저장된 디렉토리
    """
    year = str(year)
    idx = load_dashboard_index(p_type, year, output_dir)
    if idx is None:
        print(
            f"ℹ️ 저장된 이미지가 없어 즉시 표시할 수 없습니다.\n"
            f"   → show_dashboard_year_recompute({year}, final_results, io_files, "
            f"p_type='{p_type}') 를 사용해 새로 계산하며 표시하세요."
        )
        return

    from IPython.display import display, Image

    label = "경상가격" if p_type == "경상" else "불변가격"
    display(HTML(
        f"<div style='background-color:#f0fdf4;padding:12px;"
        f"border-left:6px solid #16a34a;margin-bottom:14px;font-size:15px'>"
        f"<b>📊 {year}년 EEIO 대시보드 전체 보기 ({label})</b><br>"
        f"<span style='font-size:12px;color:#6b7280'>"
        f"총 {len(idx['cards'])}개 산업 · 인트로 1장 + 산업 카드 {len(idx['cards'])}장</span></div>"
    ))

    # 1) 인트로 표지
    intro_fpath = out_path(idx["intro_file"], output_dir)
    if os.path.exists(intro_fpath):
        display(Image(filename=intro_fpath))
    else:
        print(f"⚠️ 인트로 이미지 없음: {intro_fpath}")

    # 2) 전 산업 카드 (rank 오름차순 = 총배출량 내림차순)
    cards_sorted = sorted(idx["cards"], key=lambda c: c["rank"])
    for c in cards_sorted:
        card_fpath = out_path(c["file"], output_dir)
        if os.path.exists(card_fpath):
            display(Image(filename=card_fpath))
        else:
            print(f"⚠️ 카드 이미지 없음: {card_fpath} (산업: {c['산업']})")

    print(f"✅ {year}년 {label} 대시보드 전체 ({len(cards_sorted)}개 산업) 출력 완료.")


def show_dashboard_year_recompute(
    year,
    final_results: dict,
    io_files: dict,
    p_type: str = "경상",
    exclude_keyword: str | None = "기타",
    output_dir: str = "./output",
) -> None:
    """
    ★ [연도만 입력하면 전체 출력 — 재계산 버전] ★
    해당 연도를 매번 새로 계산하여 인트로 + 전 산업 카드를 순서대로
    전부 생성/표시/저장합니다.

    plot_dashboard_with_scope() 를 그대로 호출하는 얇은 래퍼로, 위젯 없이
    "연도만 바꿔서 셀 재실행"하는 사용 패턴에 맞춘 함수입니다.
    실행 후에는 Dashboard_Index_{p_type}_{year}.json 이 저장되므로, 이후에는
    show_dashboard_year(year, p_type) 로 재계산 없이 즉시 재표시할 수 있습니다.

    Parameters
    ----------
    year            : 연도 (int 또는 str, 예: 2023)
    final_results   : run_all_years() 또는 init_pipeline() 결과의 final_c/final_k
    io_files        : {연도: IO 엑셀 경로} 딕셔너리
    p_type          : '경상' 또는 '불변'
    exclude_keyword : 제외할 산업명 (기본 '기타', 정확히 일치하는 것만 제외)
    output_dir      : 결과 저장 디렉토리
    """
    year = str(year)
    if year not in final_results:
        print(f"❌ final_results 에 {year}년 데이터가 없습니다. "
              f"보유 연도: {sorted(final_results.keys())}")
        return
    if year not in io_files:
        print(f"❌ io_files 에 {year}년 경로가 없습니다. "
              f"보유 연도: {sorted(io_files.keys())}")
        return

    plot_dashboard_with_scope(
        final_results, io_files, p_type=p_type, year=year,
        exclude_keyword=exclude_keyword, output_dir=output_dir,
    )


def show_dashboard_year_state(year, p_type: str = "경상") -> None:
    """
    ★ [연도만 입력하면 전체 출력 — init_pipeline() 사용자용 올인원 버전] ★
    init_pipeline() 을 이미 실행해 모듈 내부 _STATE 에 결과가 있는 경우,
    별도 인자 없이 연도만 넣으면 알아서:
      1) 저장된 이미지가 있으면 그것을 즉시 전체 표시
      2) 없으면 _STATE 의 final_c/final_k, io_files 를 사용해 자동으로
         재계산까지 수행하여 전체 표시
    합니다.

    사용 예시 (init_pipeline() 실행 후):
        ec.show_dashboard_year_state(2023)              # 경상가격
        ec.show_dashboard_year_state(2023, p_type="불변")
    """
    year = str(year)
    output_dir = _s("output_dir")

    idx = load_dashboard_index(p_type, year, output_dir)
    if idx is not None:
        show_dashboard_year(year, p_type=p_type, output_dir=output_dir)
        return

    print(f"ℹ️ {year}년 {p_type}가격 저장 이미지가 없어 새로 계산합니다...")
    final_results = _s("final_c") if p_type == "경상" else _s("final_k")
    io_files      = _s("io_files")
    exclude       = _s("exclude")
    show_dashboard_year_recompute(
        year, final_results, io_files,
        p_type=p_type, exclude_keyword=exclude, output_dir=output_dir,
    )


def show_saved_image(filename: str, output_dir: str = "./output") -> None:
    """output_dir 에 저장된 임의의 PNG 파일을 그대로 표시 (히트맵/막대/시계열 등 공용)."""
    fpath = out_path(filename, output_dir)
    if not os.path.exists(fpath):
        print(f"⚠️ 이미지 파일 없음: {fpath}")
        return
    from IPython.display import display, Image
    display(Image(filename=fpath))


def show_facet_timeseries_image(p_type: str = "불변", output_dir: str = "./output") -> None:
    """plot_facet_timeseries() 로 저장된 격자 시계열 이미지를 그대로 표시."""
    show_saved_image(f"Step_B_vs_BL_Timeseries_{p_type}.png", output_dir)


# ═══════════════════════════════════════════════════════════════════════════════
# 13. 기존 저장 이미지 스캔 (연도별 카드 구분)
# ═══════════════════════════════════════════════════════════════════════════════
#
#  실제 저장되는 파일명:
#    - Dashboard_00_Intro_{year}.png               (예: Dashboard_00_Intro_2018.png)
#    - Dashboard_Card_{year}_{rank:02d}_{산업명}.png (예: Dashboard_Card_2018_01_농림수산품.png)
#    - Step_B_vs_BL_Timeseries_{p_type}.png         (예: Step_B_vs_BL_Timeseries_불변.png, 연도 구분 없음)


def scan_dashboard_images(output_dir: str = "./output") -> dict:
    """
    output_dir 안의 Dashboard_00_Intro_{year}.png / Dashboard_Card_{year}_{rank}_{산업명}.png
    파일명을 스캔하여 {"years": [...], "intro": {year: path}, "cards": {year: {산업명: path}}} 반환.
    """
    import re, glob

    files = glob.glob(os.path.join(output_dir, "Dashboard_*.png"))

    intro: dict = {}
    cards: dict = {}

    for fp in files:
        fname = os.path.basename(fp)

        m_intro = re.match(r"Dashboard_00_Intro_(\d{4})\.png$", fname)
        if m_intro:
            intro[m_intro.group(1)] = fp
            continue

        m_card = re.match(r"Dashboard_Card_(\d{4})_(\d{2})_(.+)\.png$", fname)
        if m_card:
            year, safe_name = m_card.group(1), m_card.group(3)
            cards.setdefault(year, {})[safe_name] = fp
            continue

        # 구버전(연도 없는) 카드 파일도 호환 — "(연도 미상)" 그룹에 보관
        m_card_old = re.match(r"Dashboard_Card_(\d{2})_(.+)\.png$", fname)
        if m_card_old:
            safe_name = m_card_old.group(2)
            cards.setdefault("(연도 미상)", {})[safe_name] = fp

    # 인트로 표지 이미지가 실제로 존재하는 연도만 노출
    years = sorted(intro.keys())
    return {"years": years, "intro": intro, "cards": cards}


def show_dashboard_image_v2(
    year: str | None,
    sector: str | None,
    output_dir: str = "./output",
) -> None:
    """
    저장된 이미지 파일명을 스캔하여 그대로 표시. (단일 이미지 표시용, 하위 호환)

    sector=None 또는 "(전체 요약)" → 해당 연도의 인트로 표지 표시
    sector="산업명"                 → 해당 연도의 산업 카드 표시
    """
    catalog = scan_dashboard_images(output_dir)

    if sector is None or sector == "(전체 요약)":
        fpath = catalog["intro"].get(year)
        if fpath is None:
            print(f"⚠️ {year}년 인트로 이미지가 없습니다. "
                  f"사용 가능 연도: {catalog['years']}")
            return
    else:
        safe = "".join(c for c in sector if c.isalnum() or c in " _-")
        year_cards = catalog["cards"].get(year, {})
        fpath = year_cards.get(safe)
        if fpath is None:
            cand = [v for k, v in year_cards.items() if safe in k or k in safe]
            fpath = cand[0] if cand else None
        if fpath is None:
            print(f"⚠️ {year}년 '{sector}' 카드 이미지를 찾지 못했습니다. "
                  f"사용 가능 산업: {sorted(year_cards.keys())}")
            return

    from IPython.display import display, Image
    display(Image(filename=fpath))


def show_dashboard_year_v2(
    year,
    output_dir: str = "./output",
) -> None:
    """
    ★ [연도만 입력하면 전체 출력 — JSON 인덱스 없이 파일명 스캔만으로 동작하는 버전] ★

    Dashboard_Index_*.json 이 없어도, Dashboard_00_Intro_{year}.png 와
    Dashboard_Card_{year}_{rank}_{산업명}.png 파일들만 output_dir 에 있으면
    그대로 순서대로(rank 오름차순) 전부 표시합니다.

    show_dashboard_year() 와 동일하게 위젯을 사용하지 않으므로, 노트북을
    껐다 켜도 이 함수 호출 한 줄로 재현됩니다.

    Parameters
    ----------
    year       : 연도 (int 또는 str, 예: 2023)
    output_dir : 이미지가 저장된 디렉토리
    """
    import re

    year = str(year)
    catalog = scan_dashboard_images(output_dir)

    intro_fpath = catalog["intro"].get(year)
    year_cards  = catalog["cards"].get(year, {})

    if intro_fpath is None and not year_cards:
        print(f"⚠️ {year}년 저장된 대시보드 이미지가 없습니다. "
              f"사용 가능 연도: {catalog['years']}")
        return

    from IPython.display import display, Image

    display(HTML(
        f"<div style='background-color:#f0fdf4;padding:12px;"
        f"border-left:6px solid #16a34a;margin-bottom:14px;font-size:15px'>"
        f"<b>📊 {year}년 EEIO 대시보드 전체 보기</b><br>"
        f"<span style='font-size:12px;color:#6b7280'>총 {len(year_cards)}개 산업</span></div>"
    ))

    if intro_fpath and os.path.exists(intro_fpath):
        display(Image(filename=intro_fpath))
    else:
        print(f"⚠️ {year}년 인트로 이미지가 없습니다.")

    # 파일명의 rank(2자리 숫자)를 기준으로 정렬하여 순서대로 출력
    def _rank_of(fpath):
        m = re.search(r"Dashboard_Card_\d{4}_(\d{2})_", os.path.basename(fpath))
        return int(m.group(1)) if m else 999

    for safe_name, fpath in sorted(year_cards.items(), key=lambda kv: _rank_of(kv[1])):
        if os.path.exists(fpath):
            display(Image(filename=fpath))
        else:
            print(f"⚠️ 이미지 파일 없음: {fpath}")

    print(f"✅ {year}년 대시보드 전체 ({len(year_cards)}개 산업) 출력 완료.")